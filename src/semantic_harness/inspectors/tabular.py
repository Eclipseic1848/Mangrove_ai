# -*- coding: utf-8 -*-
"""常见表格文件的只读结构检查；不执行任何数据变换。"""
from __future__ import annotations

import csv
from collections import Counter
from datetime import date, datetime
from functools import lru_cache
import json
from pathlib import Path
import re
import unicodedata
import uuid
from typing import Any, Iterable, Sequence

from src.semantic_harness.inspection_models import (
    ColumnProfile,
    InspectionDiagnostic,
    InspectionStatus,
    SourceInspectionReport,
    SourceKind,
    TableProfile,
)


INSPECTOR_VERSION = "batch2-tabular-v1"
_SAMPLE_ROWS = 200
_MAX_JSON_BYTES = 10 * 1024 * 1024


def _is_empty(value: Any) -> bool:
    return value is None or value == ""


def normalize_label(value: str) -> str:
    text = unicodedata.normalize("NFKC", value or "").strip().lower()
    return re.sub(r"[\s_./·:：()（）\[\]【】\-]+", "", text)


def _safe_sample(value: Any) -> str:
    text = str(value).strip()
    if re.fullmatch(r"[^@\s]+@[^@\s]+\.[^@\s]+", text):
        left, domain = text.split("@", 1)
        return f"{left[:1]}***@{domain}"
    digits = re.sub(r"\D", "", text)
    if len(digits) >= 11:
        return f"{text[:3]}***{text[-4:]}"
    return text[:80]


def _infer_type(values: Sequence[Any]) -> str:
    non_empty = [value for value in values if not _is_empty(value)]
    if not non_empty:
        return "empty"
    kinds = set()
    for value in non_empty:
        if isinstance(value, bool):
            kinds.add("boolean")
        elif isinstance(value, int):
            kinds.add("integer")
        elif isinstance(value, float):
            kinds.add("number")
        elif isinstance(value, (datetime, date)):
            kinds.add("date")
        else:
            text = str(value).strip()
            if re.fullmatch(r"[-+]?\d+", text):
                kinds.add("integer")
            elif re.fullmatch(r"[-+]?(?:\d+\.\d+|\d+)(?:%|元|万元)?", text):
                kinds.add("number")
            elif re.fullmatch(r"\d{4}[-/.年]\d{1,2}[-/.月]\d{1,2}日?", text):
                kinds.add("date")
            else:
                kinds.add("string")
    return next(iter(kinds)) if len(kinds) == 1 else "mixed"


def _profile_table(
    *,
    artifact_id: str,
    table_index: int,
    table_name: str,
    header_row: int,
    headers: Sequence[Any],
    rows: Sequence[Sequence[Any]],
    estimated_rows: int | None,
) -> TableProfile:
    table_ref = f"artifact://{artifact_id}/table/{table_index}"
    raw_names = ["" if value is None else str(value).strip() for value in headers]
    normalized = [normalize_label(value) for value in raw_names]
    counts = Counter(item for item in normalized if item)
    columns = []
    for index, raw_name in enumerate(raw_names):
        values = [row[index] if index < len(row) else None for row in rows]
        non_empty = [value for value in values if not _is_empty(value)]
        unique = {str(value) for value in non_empty}
        duplicate_group = (
            normalized[index]
            if normalized[index] and counts[normalized[index]] > 1
            else None
        )
        samples = tuple(
            dict.fromkeys(_safe_sample(value) for value in non_empty)
        )[:8]
        columns.append(
            ColumnProfile(
                physical_ref=f"{table_ref}/column/{index}",
                artifact_id=artifact_id,
                table_ref=table_ref,
                column_index=index,
                raw_name=raw_name,
                normalized_name=normalized[index],
                inferred_type=_infer_type(values),
                null_ratio=(
                    sum(_is_empty(value) for value in values) / len(values)
                    if values
                    else 1.0
                ),
                unique_ratio=(
                    len(unique) / len(non_empty) if non_empty else 0.0
                ),
                sample_values=samples,
                duplicate_group=duplicate_group,
            )
        )
    return TableProfile(
        table_ref=table_ref,
        artifact_id=artifact_id,
        name=table_name,
        table_index=table_index,
        header_row=header_row,
        sampled_rows=len(rows),
        estimated_rows=estimated_rows,
        columns=tuple(columns),
    )


def _text_encoding(path: Path) -> str:
    """只读取前 1 MiB 探测编码，避免为了检查表头把大 CSV 全量载入内存。"""

    with path.open("rb") as handle:
        data = handle.read(1024 * 1024)
    for encoding in ("utf-8-sig", "gb18030"):
        try:
            data.decode(encoding)
            return encoding
        except UnicodeDecodeError:
            continue
    raise UnicodeDecodeError("utf-8", data, 0, min(1, len(data)), "无法识别文本编码")


def _csv_tables(path: Path, artifact_id: str) -> tuple[list[TableProfile], list[InspectionDiagnostic]]:
    encoding = _text_encoding(path)
    delimiter = "\t" if path.suffix.lower() == ".tsv" else ","
    with path.open("r", encoding=encoding, newline="") as handle:
        reader = csv.reader(handle, delimiter=delimiter)
        try:
            headers = next(reader)
        except StopIteration:
            raise ValueError("表格文件为空")
        rows = []
        for row in reader:
            if any(str(value).strip() for value in row):
                rows.append(row)
            if len(rows) >= _SAMPLE_ROWS:
                break
    diagnostics = (
        []
        if encoding == "utf-8-sig"
        else [
            InspectionDiagnostic(
                code="non_utf8_text",
                message=f"文本使用 {encoding} 解码",
            )
        ]
    )
    return [
        _profile_table(
            artifact_id=artifact_id,
            table_index=0,
            table_name=path.stem,
            header_row=1,
            headers=headers,
            rows=rows,
            estimated_rows=None,
        )
    ], diagnostics


def _xlsx_tables(path: Path, artifact_id: str) -> tuple[list[TableProfile], list[InspectionDiagnostic]]:
    from openpyxl import load_workbook

    with path.open("rb") as source:
        workbook = load_workbook(
            source,
            read_only=True,
            data_only=False,
            keep_links=False,
        )
        tables = []
        diagnostics = []
        try:
            for table_index, sheet in enumerate(workbook.worksheets):
                iterator = sheet.iter_rows(values_only=True)
                headers = None
                header_row = 1
                for row_number, row in enumerate(iterator, start=1):
                    if any(not _is_empty(value) for value in row):
                        headers = row
                        header_row = row_number
                        break
                    if row_number >= 20:
                        break
                if headers is None:
                    diagnostics.append(
                        InspectionDiagnostic(
                            code="empty_sheet",
                            message=f"工作表 {sheet.title} 没有可识别表头",
                        )
                    )
                    continue
                rows = []
                for row in iterator:
                    if any(not _is_empty(value) for value in row):
                        rows.append(row)
                    if len(rows) >= _SAMPLE_ROWS:
                        break
                tables.append(
                    _profile_table(
                        artifact_id=artifact_id,
                        table_index=table_index,
                        table_name=sheet.title,
                        header_row=header_row,
                        headers=headers,
                        rows=rows,
                        estimated_rows=max(sheet.max_row - header_row, 0),
                    )
                )
                if sheet.sheet_state != "visible":
                    diagnostics.append(
                        InspectionDiagnostic(
                            code="hidden_sheet",
                            message=f"工作表 {sheet.title} 为隐藏状态",
                        )
                    )
        finally:
            workbook.close()
    return tables, diagnostics


def _parquet_tables(path: Path, artifact_id: str) -> tuple[list[TableProfile], list[InspectionDiagnostic]]:
    import pyarrow.parquet as pq

    parquet = pq.ParquetFile(path)
    headers = parquet.schema_arrow.names
    rows: list[list[Any]] = []
    for batch in parquet.iter_batches(batch_size=_SAMPLE_ROWS):
        values = batch.to_pylist()
        rows.extend([[row.get(name) for name in headers] for row in values])
        break
    return [
        _profile_table(
            artifact_id=artifact_id,
            table_index=0,
            table_name=path.stem,
            header_row=1,
            headers=headers,
            rows=rows,
            estimated_rows=parquet.metadata.num_rows,
        )
    ], []


def _rows_from_json_value(value: Any) -> list[dict[str, Any]]:
    if isinstance(value, list):
        rows = [item for item in value if isinstance(item, dict)]
    elif isinstance(value, dict):
        if value and all(isinstance(item, dict) for item in value.values()):
            rows = list(value.values())
        else:
            rows = [value]
    else:
        rows = []
    return rows[:_SAMPLE_ROWS]


def _json_tables(path: Path, artifact_id: str) -> tuple[list[TableProfile], list[InspectionDiagnostic]]:
    if path.suffix.lower() == ".jsonl":
        rows = []
        with path.open("r", encoding="utf-8-sig") as handle:
            for line_number, line in enumerate(handle, start=1):
                if not line.strip():
                    continue
                value = json.loads(line)
                if not isinstance(value, dict):
                    raise ValueError(f"JSONL 第 {line_number} 行不是对象")
                rows.append(value)
                if len(rows) >= _SAMPLE_ROWS:
                    break
    else:
        if path.stat().st_size > _MAX_JSON_BYTES:
            raise OverflowError("JSON 数组超过检查上限，请改用 JSONL")
        value = json.loads(path.read_text(encoding="utf-8-sig"))
        rows = _rows_from_json_value(value)
    if not rows:
        raise ValueError("未发现对象记录")
    headers = list(dict.fromkeys(key for row in rows for key in row))
    matrix = [[row.get(name) for name in headers] for row in rows]
    return [
        _profile_table(
            artifact_id=artifact_id,
            table_index=0,
            table_name=path.stem,
            header_row=1,
            headers=headers,
            rows=matrix,
            estimated_rows=None,
        )
    ], []


def _detected_format(path: Path) -> str:
    extension = path.suffix.lower().lstrip(".")
    try:
        import filetype

        kind = filetype.guess(str(path))
        if kind:
            return kind.extension
    except Exception:
        pass
    try:
        label = _magika().identify_path(path).output.label
        if label not in {"txt", "unknown", "undefined"}:
            return label
    except Exception:
        pass
    return extension


@lru_cache(maxsize=1)
def _magika():
    """延迟加载成熟内容识别模型，避免二进制格式命中 filetype 时支付开销。"""

    from magika import Magika

    return Magika()


def inspect_tabular_path(
    *,
    artifact_id: str,
    artifact_sha256: str,
    path: Path,
    original_name: str,
    declared_media_type: str,
) -> SourceInspectionReport:
    """只读检查一个表格文件，失败时返回分类报告而非泄漏异常堆栈。"""

    suffix = Path(original_name).suffix.lower()
    detected = _detected_format(path)
    status = InspectionStatus.READY
    tables: list[TableProfile] = []
    diagnostics: list[InspectionDiagnostic] = []
    try:
        if suffix in {".csv", ".tsv"}:
            tables, diagnostics = _csv_tables(path, artifact_id)
        elif suffix == ".xlsx":
            tables, diagnostics = _xlsx_tables(path, artifact_id)
        elif suffix == ".parquet":
            tables, diagnostics = _parquet_tables(path, artifact_id)
        elif suffix in {".json", ".jsonl"}:
            tables, diagnostics = _json_tables(path, artifact_id)
        else:
            status = InspectionStatus.UNSUPPORTED
            diagnostics.append(
                InspectionDiagnostic(
                    code="unsupported_tabular_format",
                    message=f"当前不支持表格格式 {suffix or 'unknown'}",
                )
            )
        if status == InspectionStatus.READY and not tables:
            status = InspectionStatus.CORRUPT
            diagnostics.append(
                InspectionDiagnostic(
                    code="no_tables",
                    message="没有发现可检查的表格",
                )
            )
    except OverflowError as exc:
        status = InspectionStatus.OVER_LIMIT
        diagnostics.append(
            InspectionDiagnostic(code="over_limit", message=str(exc))
        )
    except Exception as exc:
        text = str(exc).lower()
        status = (
            InspectionStatus.ENCRYPTED
            if "password" in text or "encrypt" in text
            else InspectionStatus.CORRUPT
        )
        diagnostics.append(
            InspectionDiagnostic(
                code="encrypted" if status == InspectionStatus.ENCRYPTED else "parse_failed",
                message=f"{type(exc).__name__}: {str(exc)[:300]}",
            )
        )
    return SourceInspectionReport(
        inspection_id=f"insp_{uuid.uuid4().hex[:16]}",
        inspector_version=INSPECTOR_VERSION,
        artifact_id=artifact_id,
        artifact_sha256=artifact_sha256,
        size_bytes=path.stat().st_size,
        original_name=original_name,
        declared_media_type=declared_media_type,
        detected_format=detected or suffix.lstrip(".") or "unknown",
        source_kind=SourceKind.TABULAR,
        status=status,
        tables=tuple(tables),
        diagnostics=tuple(diagnostics),
    )
