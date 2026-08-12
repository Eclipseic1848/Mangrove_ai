# -*- coding: utf-8 -*-
"""DuckDB 唯一业务执行引擎：只执行结构化 PhysicalPlan。"""
from __future__ import annotations

import csv
from dataclasses import dataclass
import hashlib
import json
from pathlib import Path
import re
import time
from typing import Any, Mapping, Sequence
import uuid

import pyarrow as pa
import pyarrow.compute as pc
import pyarrow.parquet as pq
import sqlglot
from sqlglot import exp

from .models import (
    ArtifactRef,
    ExecutionLedger,
    FailureKind,
    LineageEvent,
    PredicateOperator,
    ResourceUsage,
    ToolResult,
    ToolStatus,
)
from .physical_models import (
    AggregateFunction,
    AggregateStep,
    DeduplicateStep,
    FilterStep,
    JoinCardinality,
    JoinStep,
    PhysicalPlan,
    PhysicalPlanStatus,
    ProjectStep,
    RenameStep,
    SortStep,
    UnionStep,
)


_LINEAGE_COLUMN = "__mg_lineage"
_SOURCE_ROW_COLUMN = "__mg_source_row_id"
_SOURCE_NUMBER_COLUMN = "__mg_source_row_number"


@dataclass(frozen=True)
class ExecutionBundle:
    tool_result: ToolResult
    result_path: Path | None
    lineage_path: Path | None
    output_table: pa.Table | None
    evidence_rows: tuple[dict[str, Any], ...]


def _sha256_path(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _qid(value: str) -> str:
    return exp.to_identifier(value, quoted=True).sql(dialect="duckdb")


def _source_row_id(artifact_hash: str, table_ref: str, row_number: int) -> str:
    raw = f"{artifact_hash}|{table_ref}|{row_number}".encode("utf-8")
    return hashlib.sha256(raw).hexdigest()


def _read_text_rows(
    path: Path,
    *,
    delimiter: str,
    header_row: int,
) -> list[tuple[int, list[Any]]]:
    data = path.read_bytes()
    encoding = "utf-8-sig"
    try:
        text = data.decode(encoding)
    except UnicodeDecodeError:
        encoding = "gb18030"
        text = data.decode(encoding)
    reader = csv.reader(text.splitlines(), delimiter=delimiter)
    rows = list(reader)
    return [
        (row_number, row)
        for row_number, row in enumerate(
            rows[header_row:],
            start=header_row + 1,
        )
        if any(str(value).strip() for value in row)
    ]


def _read_xlsx_rows(
    path: Path,
    *,
    table_index: int,
    header_row: int,
) -> list[tuple[int, list[Any]]]:
    from openpyxl import load_workbook

    with path.open("rb") as formula_source, path.open("rb") as value_source:
        formulas = load_workbook(
            formula_source,
            read_only=True,
            data_only=False,
            keep_links=False,
        )
        values = load_workbook(
            value_source,
            read_only=True,
            data_only=True,
            keep_links=False,
        )
        try:
            formula_sheet = formulas.worksheets[table_index]
            value_sheet = values.worksheets[table_index]
            output: list[tuple[int, list[Any]]] = []
            for row_number, (formula_row, value_row) in enumerate(
                zip(
                    formula_sheet.iter_rows(values_only=False),
                    value_sheet.iter_rows(values_only=True),
                ),
                start=1,
            ):
                if row_number <= header_row:
                    continue
                resolved = []
                for formula_cell, cached_value in zip(
                    formula_row,
                    value_row,
                ):
                    if (
                        formula_cell.data_type == "f"
                        and cached_value is None
                    ):
                        raise ValueError(
                            f"工作表 {formula_sheet.title} "
                            f"第 {row_number} 行包含无缓存公式值"
                        )
                    resolved.append(cached_value)
                if any(value not in (None, "") for value in resolved):
                    output.append((row_number, resolved))
            return output
        finally:
            formulas.close()
            values.close()


def _read_polars_rows(
    path: Path,
    detected_format: str,
    *,
    header_row: int,
) -> list[tuple[int, list[Any]]]:
    import polars as pl

    if detected_format == "parquet":
        frame = pl.read_parquet(path)
    elif detected_format == "jsonl":
        frame = pl.read_ndjson(path)
    elif detected_format == "json":
        frame = pl.read_json(path)
    else:
        raise ValueError(f"不支持 Polars 读取格式：{detected_format}")
    return list(
        enumerate(frame.rows(), start=header_row + 1)
    )


def _load_source(source, path: Path) -> tuple[pa.Table, dict[str, dict[str, Any]]]:
    """按检查报告的列序号读取，不用原始重复表头做业务寻址。"""

    if _sha256_path(path) != source.artifact_sha256:
        raise ValueError(f"来源哈希已变化：{source.artifact_id}")
    if source.detected_format == "csv":
        rows = _read_text_rows(path, delimiter=",", header_row=source.header_row)
    elif source.detected_format == "tsv":
        rows = _read_text_rows(path, delimiter="\t", header_row=source.header_row)
    elif source.detected_format == "xlsx":
        rows = _read_xlsx_rows(
            path,
            table_index=source.table_index,
            header_row=source.header_row,
        )
    else:
        rows = _read_polars_rows(
            path,
            source.detected_format,
            header_row=source.header_row,
        )

    data: dict[str, list[Any]] = {
        column.output_name: [] for column in source.columns
    }
    row_numbers: list[int] = []
    row_ids: list[str] = []
    evidence: dict[str, dict[str, Any]] = {}
    for row_number, row in rows:
        row_id = _source_row_id(
            source.artifact_sha256, source.table_ref, row_number
        )
        values: dict[str, Any] = {}
        for column in source.columns:
            value = row[column.column_index] if column.column_index < len(row) else None
            if value not in (None, "") and isinstance(value, str):
                text = value.strip()
                if column.inferred_type == "integer":
                    value = int(text)
                elif column.inferred_type == "number":
                    multiplier = 10_000 if text.endswith("万元") else 1
                    percentage = text.endswith("%")
                    numeric = re.sub(r"(万元|元|%)$", "", text)
                    value = float(numeric) * multiplier
                    if percentage:
                        value /= 100
            data[column.output_name].append(value)
            values[column.output_name] = value
        row_numbers.append(row_number)
        row_ids.append(row_id)
        evidence[row_id] = {
            "source_row_id": row_id,
            "artifact_id": source.artifact_id,
            "table_ref": source.table_ref,
            "row_number": row_number,
            "values": values,
        }
    arrays = {name: pa.array(values) for name, values in data.items()}
    arrays[_SOURCE_NUMBER_COLUMN] = pa.array(row_numbers, type=pa.int64())
    arrays[_SOURCE_ROW_COLUMN] = pa.array(row_ids, type=pa.string())
    arrays[_LINEAGE_COLUMN] = pa.array(
        [[row_id] for row_id in row_ids],
        type=pa.list_(pa.string()),
    )
    return pa.table(arrays), evidence


def _placeholder(value: Any, params: list[Any]) -> str:
    params.append(value)
    return "?"


def _condition_sql(condition, params: list[Any]) -> str:
    column = _qid(condition.column)
    operator = condition.operator
    if operator == PredicateOperator.IS_NULL:
        return f"{column} IS NULL"
    if operator == PredicateOperator.NOT_NULL:
        return f"{column} IS NOT NULL"
    if operator in {PredicateOperator.IN, PredicateOperator.NOT_IN}:
        placeholders = ", ".join(
            _placeholder(value, params) for value in condition.values
        )
        keyword = "IN" if operator == PredicateOperator.IN else "NOT IN"
        return f"{column} {keyword} ({placeholders})"
    placeholder = _placeholder(condition.value, params)
    binary = {
        PredicateOperator.EQ: "=",
        PredicateOperator.NE: "<>",
        PredicateOperator.GT: ">",
        PredicateOperator.GTE: ">=",
        PredicateOperator.LT: "<",
        PredicateOperator.LTE: "<=",
    }
    if operator in binary:
        if (
            operator in {PredicateOperator.EQ, PredicateOperator.NE}
            and not condition.case_sensitive
            and isinstance(condition.value, str)
        ):
            return f"lower(CAST({column} AS VARCHAR)) {binary[operator]} lower({placeholder})"
        return f"{column} {binary[operator]} {placeholder}"
    if operator == PredicateOperator.CONTAINS:
        if condition.case_sensitive:
            return f"contains(CAST({column} AS VARCHAR), {placeholder})"
        return (
            f"contains(lower(CAST({column} AS VARCHAR)), lower({placeholder}))"
        )
    if operator == PredicateOperator.REGEX:
        return f"regexp_matches(CAST({column} AS VARCHAR), {placeholder})"
    raise ValueError(f"不支持筛选操作符：{operator.value}")


def _order_sql(keys: Sequence[Any]) -> str:
    parts = []
    for key in keys:
        nulls = "NULLS LAST" if key.nulls_last else "NULLS FIRST"
        parts.append(
            f"{_qid(key.column)} {key.direction.value.upper()} {nulls}"
        )
    return ", ".join(parts)


def _validate_select_sql(query: str) -> None:
    """用 SQLGlot 复核生成 SQL；任何外部读取、DDL、命令节点均拒绝。"""

    tree = sqlglot.parse_one(query, read="duckdb")
    forbidden = (
        exp.Command,
        exp.Create,
        exp.Delete,
        exp.Drop,
        exp.Insert,
        exp.Update,
        exp.Copy,
    )
    if not isinstance(tree, (exp.Select, exp.Union, exp.Subquery)):
        raise ValueError("物理编译器只允许 SELECT")
    if any(tree.find(node) is not None for node in forbidden):
        raise ValueError("生成 SQL 含禁止节点")
    for function in tree.find_all(exp.Anonymous):
        if function.name.lower() in {
            "read_csv",
            "read_csv_auto",
            "read_json",
            "read_parquet",
            "httpfs",
        }:
            raise ValueError("禁止 SQL 直接读取路径或网络")
    if any(
        type(node).__name__.lower()
        in {"readparquet", "readcsv", "readjson"}
        for node in tree.walk()
    ):
        raise ValueError("禁止 SQL 直接读取路径或网络")


def _compile_queries(plan: PhysicalPlan) -> tuple[dict[str, str], list[Any]]:
    queries = {
        source.source_id: f"SELECT * FROM {_qid(source.source_id)}"
        for source in plan.sources
    }
    params: list[Any] = []
    for step in plan.steps:
        inputs = [f"({queries[item]})" for item in step.input_ids]
        if isinstance(step, UnionStep):
            query = " UNION ALL BY NAME ".join(
                f"SELECT * FROM {item}" for item in inputs
            )
        elif isinstance(step, FilterStep):
            conditions = " AND ".join(
                _condition_sql(item, params) for item in step.conditions
            )
            query = f"SELECT * FROM {inputs[0]} AS input WHERE {conditions}"
        elif isinstance(step, ProjectStep):
            columns = ", ".join(
                f"{_qid(item.source)} AS {_qid(item.output)}"
                for item in step.columns
            )
            query = (
                f"SELECT {columns}, {_qid(_LINEAGE_COLUMN)} "
                f"FROM {inputs[0]} AS input"
            )
        elif isinstance(step, RenameStep):
            replacements = ", ".join(
                f"{_qid(source)} AS {_qid(output)}"
                for source, output in step.mapping.items()
            )
            query = (
                f"SELECT * RENAME ({replacements}) "
                f"FROM {inputs[0]} AS input"
            )
        elif isinstance(step, SortStep):
            query = (
                f"SELECT * FROM {inputs[0]} AS input "
                f"ORDER BY {_order_sql(step.keys)}"
            )
        elif isinstance(step, DeduplicateStep):
            partition = ", ".join(_qid(item) for item in step.keys)
            order = (
                _order_sql(step.order_by)
                if step.order_by
                else ", ".join(_qid(item) for item in step.keys)
            )
            using_keys = ", ".join(_qid(item) for item in step.keys)
            query = (
                f"WITH source AS (SELECT * FROM {inputs[0]} AS input), "
                "chosen AS ("
                f"SELECT * EXCLUDE (__mg_rank) FROM (SELECT *, row_number() "
                f"OVER (PARTITION BY {partition} ORDER BY {order}) AS __mg_rank "
                "FROM source) ranked WHERE __mg_rank = 1"
                "), evidence AS ("
                f"SELECT {partition}, "
                f"list_sort(list_distinct(flatten(list({_qid(_LINEAGE_COLUMN)})))) "
                f"AS {_qid('__mg_all_lineage')} FROM source GROUP BY {partition}"
                ") SELECT chosen.* EXCLUDE "
                f"({_qid(_LINEAGE_COLUMN)}), evidence.{_qid('__mg_all_lineage')} "
                f"AS {_qid(_LINEAGE_COLUMN)} FROM chosen JOIN evidence "
                f"USING ({using_keys})"
            )
        elif isinstance(step, AggregateStep):
            groups = ", ".join(_qid(item) for item in step.group_by)
            expressions = []
            for item in step.aggregates:
                if item.function == AggregateFunction.COUNT:
                    value = "count(*)"
                else:
                    value = f"{item.function.value}({_qid(item.column)})"
                expressions.append(f"{value} AS {_qid(item.output)}")
            expressions.append(
                f"list_sort(list_distinct(flatten(list({_qid(_LINEAGE_COLUMN)})))) "
                f"AS {_qid(_LINEAGE_COLUMN)}"
            )
            select = ", ".join(
                [*( [groups] if groups else []), *expressions]
            )
            group_clause = f" GROUP BY {groups}" if groups else ""
            query = f"SELECT {select} FROM {inputs[0]} AS input{group_clause}"
        elif isinstance(step, JoinStep):
            key_condition = " AND ".join(
                f"l.{_qid(key.left)} = r.{_qid(key.right)}"
                for key in step.keys
            )
            right_excludes = ", ".join(
                [_LINEAGE_COLUMN, *[key.right for key in step.keys]]
            )
            query = (
                f"SELECT l.* EXCLUDE ({_qid(_LINEAGE_COLUMN)}), "
                f"r.* EXCLUDE ({', '.join(_qid(item) for item in right_excludes)}), "
                f"list_sort(list_distinct(list_concat(l.{_qid(_LINEAGE_COLUMN)}, "
                f"r.{_qid(_LINEAGE_COLUMN)}))) AS {_qid(_LINEAGE_COLUMN)} "
                f"FROM {inputs[0]} AS l {step.join_kind.value.upper()} JOIN "
                f"{inputs[1]} AS r ON {key_condition}"
            )
        else:  # pragma: no cover - Pydantic 判别联合保证封闭
            raise ValueError(f"未知步骤：{type(step).__name__}")
        _validate_select_sql(query)
        queries[step.step_id] = query
    return queries, params


def _check_join_cardinality(
    connection,
    plan: PhysicalPlan,
    queries: Mapping[str, str],
    params: Sequence[Any],
) -> None:
    for step in plan.steps:
        if not isinstance(step, JoinStep):
            continue
        for side, input_id, keys in (
            ("left", step.input_ids[0], tuple(key.left for key in step.keys)),
            ("right", step.input_ids[1], tuple(key.right for key in step.keys)),
        ):
            must_unique = (
                step.cardinality == JoinCardinality.ONE_TO_ONE
                or (
                    step.cardinality == JoinCardinality.MANY_TO_ONE
                    and side == "right"
                )
                or (
                    step.cardinality == JoinCardinality.ONE_TO_MANY
                    and side == "left"
                )
            )
            if not must_unique:
                continue
            columns = ", ".join(_qid(item) for item in keys)
            query = (
                "SELECT count(*) FROM ("
                f"SELECT {columns}, count(*) AS n FROM ({queries[input_id]}) "
                f"GROUP BY {columns} HAVING count(*) > 1"
                ") duplicates"
            )
            input_query = queries[input_id]
            used_params = list(params[: input_query.count("?")])
            if connection.execute(query, used_params).fetchone()[0]:
                raise ValueError(
                    f"join_cardinality_violation：{side} 连接键不唯一"
                )


def _check_union_schemas(
    connection,
    plan: PhysicalPlan,
    queries: Mapping[str, str],
    params: Sequence[Any],
) -> None:
    """禁止 UNION ALL BY NAME 自动补空列或隐式改型。"""

    for step in plan.steps:
        if not isinstance(step, UnionStep):
            continue
        schemas = []
        for input_id in step.input_ids:
            query = queries[input_id]
            used_params = list(params[: query.count("?")])
            rows = connection.execute(
                f"DESCRIBE {query}",
                used_params,
            ).fetchall()
            schemas.append(
                tuple(
                    (row[0], row[1])
                    for row in rows
                    if not str(row[0]).startswith("__mg_")
                )
            )
        if any(schema != schemas[0] for schema in schemas[1:]):
            raise ValueError(
                "union_schema_conflict：来源列集合、顺序或类型不一致"
            )


def _artifact_ref(path: Path, *, artifact_id: str, kind: str) -> ArtifactRef:
    return ArtifactRef(
        artifact_id=artifact_id,
        kind=kind,
        media_type="application/vnd.apache.parquet",
        sha256=_sha256_path(path),
        size_bytes=path.stat().st_size,
    )


def _numeric_totals(table: pa.Table) -> dict[str, str]:
    totals = {}
    for field in table.schema:
        if field.name.startswith("__mg_"):
            continue
        if not (
            pa.types.is_integer(field.type)
            or pa.types.is_floating(field.type)
            or pa.types.is_decimal(field.type)
        ):
            continue
        value = pc.sum(table[field.name]).as_py()
        totals[field.name] = str(value) if value is not None else "0"
    return totals


def execute_physical_plan(
    plan: PhysicalPlan,
    *,
    artifact_paths: Mapping[str, Path],
    output_dir: Path,
) -> ExecutionBundle:
    """执行已冻结计划；失败即停，不切换 Polars/Pandas 业务语义。"""

    started = time.perf_counter()
    call_id = f"call_{uuid.uuid4().hex[:16]}"
    if plan.status != PhysicalPlanStatus.READY:
        return ExecutionBundle(
            tool_result=ToolResult(
                call_id=call_id,
                capability_id=plan.capability_id,
                capability_version=plan.capability_version,
                status=ToolStatus.NEEDS_INPUT,
                failure_kind=FailureKind.NEEDS_USER,
                error_message="；".join(plan.diagnostics),
                resource_usage=ResourceUsage(duration_ms=0),
            ),
            result_path=None,
            lineage_path=None,
            output_table=None,
            evidence_rows=(),
        )

    output_dir.mkdir(parents=True, exist_ok=True)
    source_tables: dict[str, pa.Table] = {}
    evidence_index: dict[str, dict[str, Any]] = {}
    input_bytes = 0
    seen_artifacts: set[str] = set()
    try:
        for source in plan.sources:
            path = Path(artifact_paths[source.artifact_id]).resolve()
            if not path.is_file():
                raise ValueError(f"来源文件不存在：{source.artifact_id}")
            if source.artifact_id not in seen_artifacts:
                input_bytes += path.stat().st_size
                seen_artifacts.add(source.artifact_id)
            if input_bytes > plan.runtime_policy.max_input_bytes:
                raise OverflowError("输入字节数超过 PhysicalPlan 资源上限")
            table, evidence = _load_source(source, path)
            source_tables[source.source_id] = table
            evidence_index.update(evidence)
            if (
                sum(item.num_rows for item in source_tables.values())
                > plan.runtime_policy.max_input_rows
            ):
                raise OverflowError("输入行数超过 PhysicalPlan 资源上限")

        # 严格拒绝跨来源同名字段的类型漂移，禁止自动 cast/null-fill。
        schemas: dict[str, pa.DataType] = {}
        for table in source_tables.values():
            for field in table.schema:
                if field.name.startswith("__mg_"):
                    continue
                previous = schemas.get(field.name)
                if previous is not None and previous != field.type:
                    raise ValueError(
                        f"type_conflict：{field.name} 同时出现 {previous} 与 {field.type}"
                    )
                schemas[field.name] = field.type

        import duckdb

        connection = duckdb.connect(database=":memory:")
        try:
            connection.execute(
                f"SET threads = {plan.runtime_policy.threads}"
            )
            connection.execute(
                f"SET memory_limit = '{plan.runtime_policy.memory_limit}'"
            )
            connection.execute("SET enable_external_access = false")
            connection.execute("SET autoload_known_extensions = false")
            connection.execute("SET autoinstall_known_extensions = false")
            connection.execute("SET allow_community_extensions = false")
            for source_id, table in source_tables.items():
                connection.register(source_id, table)
            connection.execute("SET lock_configuration = true")
            queries, params = _compile_queries(plan)
            _check_union_schemas(connection, plan, queries, params)
            _check_join_cardinality(connection, plan, queries, params)
            final_query = queries[plan.final_step_id]
            output_with_lineage = connection.execute(
                final_query, params
            ).fetch_arrow_table()
        finally:
            connection.close()

        lineage_values = output_with_lineage[_LINEAGE_COLUMN].to_pylist()
        visible_table = output_with_lineage.drop([_LINEAGE_COLUMN])
        if tuple(visible_table.column_names) != plan.visible_columns:
            raise ValueError(
                "visible_column_mismatch：实际列与物理计划声明不一致"
            )

        run_id = uuid.uuid4().hex
        output_ids = []
        lineage_rows = []
        covered = set()
        for index, source_ids in enumerate(lineage_values):
            source_ids = tuple(dict.fromkeys(source_ids or ()))
            raw = f"{plan.physical_plan_id}|{index}|{'|'.join(source_ids)}"
            output_id = hashlib.sha256(raw.encode("utf-8")).hexdigest()
            output_ids.append(output_id)
            for source_id in source_ids:
                covered.add(source_id)
                evidence = evidence_index[source_id]
                lineage_rows.append(
                    {
                        "output_record_id": output_id,
                        "source_row_id": source_id,
                        "artifact_id": evidence["artifact_id"],
                        "table_ref": evidence["table_ref"],
                        "row_number": evidence["row_number"],
                        "evidence_json": json.dumps(
                            evidence["values"],
                            ensure_ascii=False,
                            sort_keys=True,
                            default=str,
                        ),
                    }
                )
        visible_table = visible_table.append_column(
            "__mg_output_record_id", pa.array(output_ids, type=pa.string())
        )
        result_path = output_dir / f"{run_id}-result.parquet"
        lineage_path = output_dir / f"{run_id}-lineage.parquet"
        pq.write_table(visible_table, result_path, compression="zstd")
        lineage_table = pa.Table.from_pylist(
            lineage_rows,
            schema=pa.schema(
                [
                    ("output_record_id", pa.string()),
                    ("source_row_id", pa.string()),
                    ("artifact_id", pa.string()),
                    ("table_ref", pa.string()),
                    ("row_number", pa.int64()),
                    ("evidence_json", pa.string()),
                ]
            ),
        )
        pq.write_table(lineage_table, lineage_path, compression="zstd")

        output_ref = _artifact_ref(
            result_path,
            artifact_id=f"result_{run_id}",
            kind="tabular_result",
        )
        lineage_ref = _artifact_ref(
            lineage_path,
            artifact_id=f"lineage_{run_id}",
            kind="record_lineage",
        )
        tool_result = ToolResult(
            call_id=call_id,
            capability_id=plan.capability_id,
            capability_version=plan.capability_version,
            status=ToolStatus.SUCCEEDED,
            input_artifacts=tuple(
                ArtifactRef(
                    artifact_id=source.artifact_id,
                    kind="uploaded_source",
                    media_type="application/octet-stream",
                    sha256=source.artifact_sha256,
                    size_bytes=Path(
                        artifact_paths[source.artifact_id]
                    ).stat().st_size,
                )
                for source in {
                    item.artifact_id: item for item in plan.sources
                }.values()
            ),
            output_artifacts=(output_ref, lineage_ref),
            ledger=ExecutionLedger(
                input_records=sum(table.num_rows for table in source_tables.values()),
                output_records=visible_table.num_rows,
                filtered_out_records=max(
                    sum(table.num_rows for table in source_tables.values())
                    - visible_table.num_rows,
                    0,
                ),
                input_bytes=input_bytes,
                output_bytes=result_path.stat().st_size + lineage_path.stat().st_size,
            ),
            lineage=(
                LineageEvent(
                    event="record_lineage_materialized",
                    input_artifact_ids=tuple(
                        source.artifact_id for source in plan.sources
                    ),
                    output_artifact_ids=(
                        output_ref.artifact_id,
                        lineage_ref.artifact_id,
                    ),
                    details={
                        "output_records": visible_table.num_rows,
                        "lineage_rows": len(lineage_rows),
                        "covered_source_rows": len(covered),
                    },
                ),
            ),
            facts={
                "visible_columns": list(plan.visible_columns),
                "table_count": 1,
                "lineage_coverage": (
                    sum(bool(item) for item in lineage_values)
                    / len(lineage_values)
                    if lineage_values
                    else 1.0
                ),
                "reconciliation": {
                    "input_rows_by_source": {
                        source_id: table.num_rows
                        for source_id, table in source_tables.items()
                    },
                    "output_rows": visible_table.num_rows,
                    "input_numeric_totals_by_source": {
                        source_id: _numeric_totals(table)
                        for source_id, table in source_tables.items()
                    },
                    "output_numeric_totals": _numeric_totals(visible_table),
                    "grain_changed": any(
                        isinstance(step, (JoinStep, AggregateStep))
                        for step in plan.steps
                    ),
                },
            },
            tool_config_summary={
                "runtime_profile": plan.runtime_policy.profile.value,
                "threads": plan.runtime_policy.threads,
                "memory_limit": plan.runtime_policy.memory_limit,
                "external_access": False,
                "configuration_locked": True,
                "business_engine": "duckdb",
                "loader": "polars/openpyxl",
            },
            resource_usage=ResourceUsage(
                duration_ms=int((time.perf_counter() - started) * 1000)
            ),
        )
        return ExecutionBundle(
            tool_result=tool_result,
            result_path=result_path,
            lineage_path=lineage_path,
            output_table=visible_table,
            evidence_rows=tuple(lineage_rows),
        )
    except Exception as exc:
        message = f"{type(exc).__name__}: {str(exc)}"
        needs_user = any(
            marker in str(exc)
            for marker in (
                "type_conflict",
                "join_cardinality_violation",
                "无缓存公式值",
                "visible_column_mismatch",
                "union_schema_conflict",
            )
        )
        return ExecutionBundle(
            tool_result=ToolResult(
                call_id=call_id,
                capability_id=plan.capability_id,
                capability_version=plan.capability_version,
                status=(
                    ToolStatus.NEEDS_INPUT if needs_user else ToolStatus.FAILED
                ),
                failure_kind=(
                    FailureKind.NEEDS_USER
                    if needs_user
                    else FailureKind.INVALID_PLAN
                ),
                error_message=message[:1000],
                tool_config_summary={
                    "business_engine": "duckdb",
                    "fallback_used": False,
                },
                resource_usage=ResourceUsage(
                    duration_ms=int((time.perf_counter() - started) * 1000)
                ),
            ),
            result_path=None,
            lineage_path=None,
            output_table=None,
            evidence_rows=(),
        )
