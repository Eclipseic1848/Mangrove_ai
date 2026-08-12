# -*- coding: utf-8 -*-
"""结构化表格解析器（Phase 2 Task 4）。

使用成熟库，不手搓解析：
- CSV/TSV：标准库 csv 模块，逐行流式，坏行隔离
- Excel：openpyxl read_only 模式，逐 Sheet 逐行
- Parquet：pyarrow ParquetFile.iter_batches，按 row group 流式

坏行/坏记录进入 parse rejects，保留位置（行号/Sheet/row group）。
parse_stream 逐批产出，支持大文件流式。
"""
from __future__ import annotations

import csv
import hashlib
import io
import json
import logging
from typing import Any, Dict, Iterator, List, Optional, Tuple

from src.data_prep.models import RawArtifact, RecordEnvelope

from .registry import Parser

logger = logging.getLogger(__name__)

_XLSX_MEDIA = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
_PARQUET_MEDIA = "application/vnd.apache.parquet"


class TabularParser(Parser):
    """CSV/TSV/Excel/Parquet 解析器。用成熟库流式读取。"""

    name = "tabular"
    media_types = ("text/csv", "text/tab-separated-values", "text/plain", _XLSX_MEDIA, _PARQUET_MEDIA)
    extensions = ("csv", "tsv", "xlsx", "parquet")

    # ------------------------------------------------------------------
    # 格式识别
    # ------------------------------------------------------------------
    def _format(self, artifact: RawArtifact) -> str:
        uri = artifact.uri or artifact.storage_path or ""
        ext = uri.rsplit(".", 1)[-1].lower() if "." in uri else ""
        if ext == "xlsx":
            return "xlsx"
        if ext == "parquet":
            return "parquet"
        return "csv"  # csv/tsv 统一走 csv 模块

    def _delimiter(self, artifact: RawArtifact) -> str:
        uri = artifact.uri or artifact.storage_path or ""
        ext = uri.rsplit(".", 1)[-1].lower() if "." in uri else "csv"
        return "\t" if ext == "tsv" else ","

    # ------------------------------------------------------------------
    # 统一记录构造
    # ------------------------------------------------------------------
    def _make_record(
        self,
        artifact: RawArtifact,
        row: Dict[str, Any],
        *,
        parser_name: str,
        position: Dict[str, Any],
    ) -> RecordEnvelope:
        content_hash = hashlib.sha256(
            json.dumps(row, ensure_ascii=False, sort_keys=True, default=str).encode("utf-8")
        ).hexdigest()
        rid = hashlib.sha256(
            f"{artifact.artifact_id}:{parser_name}:{position}".encode("utf-8")
        ).hexdigest()[:16]
        return RecordEnvelope(
            record_id=rid,
            data=dict(row),
            meta={
                "source_id": artifact.source_id,
                "artifact_id": artifact.artifact_id,
                "parser": parser_name,
                "position": position,
                "content_hash": content_hash,
            },
        )

    # ------------------------------------------------------------------
    # CSV/TSV 迭代
    # ------------------------------------------------------------------
    def _decode(self, raw_bytes: bytes) -> str:
        for enc in ("utf-8-sig", "gbk"):
            try:
                return raw_bytes.decode(enc)
            except UnicodeDecodeError:
                continue
        return raw_bytes.decode("utf-8", errors="replace")

    def _iter_csv(
        self, artifact: RawArtifact, raw_bytes: bytes
    ) -> Iterator[Tuple[Optional[Dict[str, Any]], Optional[Dict], Optional[str]]]:
        """yield (row_dict, None, None) 或 (None, None, err) 表示坏行。"""
        text = self._decode(raw_bytes)
        reader = csv.reader(io.StringIO(text), delimiter=self._delimiter(artifact))
        try:
            header = next(reader)
        except StopIteration:
            return
        for row_no, row in enumerate(reader, start=2):
            if not row:
                continue
            if len(row) != len(header):
                yield None, {"position": {"row": row_no}}, "csv_bad_row"
                continue
            yield dict(zip(header, row)), {"position": {"row": row_no}}, None

    # ------------------------------------------------------------------
    # Excel 迭代（openpyxl read_only）
    # ------------------------------------------------------------------
    def _iter_excel(
        self, artifact: RawArtifact, raw_bytes: bytes
    ) -> Iterator[Tuple[Optional[Dict[str, Any]], Optional[Dict], Optional[str]]]:
        from openpyxl import load_workbook

        wb = load_workbook(io.BytesIO(raw_bytes), read_only=True, data_only=True)
        try:
            for ws in wb.worksheets:
                header: Optional[List[str]] = None
                for row_no, row in enumerate(ws.iter_rows(values_only=True), start=1):
                    if header is None:
                        header = [("" if c is None else str(c)) for c in row]
                        continue
                    if not any(c is not None for c in row):
                        continue
                    data = {}
                    for i, h in enumerate(header):
                        data[h] = row[i] if i < len(row) else None
                    yield data, {"position": {"row": row_no, "sheet": ws.title}}, None
        finally:
            wb.close()

    # ------------------------------------------------------------------
    # Parquet 迭代（pyarrow iter_batches）
    # ------------------------------------------------------------------
    def _iter_parquet(
        self, artifact: RawArtifact, raw_bytes: bytes
    ) -> Iterator[Tuple[Optional[Dict[str, Any]], Optional[Dict], Optional[str]]]:
        import pyarrow.parquet as pq

        pf = pq.ParquetFile(io.BytesIO(raw_bytes))
        global_row = 0
        for batch in pf.iter_batches(batch_size=10_000):
            for record in batch.to_pylist():
                global_row += 1
                yield dict(record), {"position": {"row": global_row}}, None

    def _iter_items(
        self, artifact: RawArtifact, raw_bytes: bytes
    ) -> Iterator[Tuple[Optional[Dict[str, Any]], Optional[Dict], Optional[str]]]:
        fmt = self._format(artifact)
        if fmt == "xlsx":
            yield from self._iter_excel(artifact, raw_bytes)
        elif fmt == "parquet":
            yield from self._iter_parquet(artifact, raw_bytes)
        else:
            yield from self._iter_csv(artifact, raw_bytes)

    def _parser_name(self, artifact: RawArtifact) -> str:
        fmt = self._format(artifact)
        return {"xlsx": "tabular_xlsx", "parquet": "tabular_parquet"}.get(fmt, "tabular_csv")

    # ------------------------------------------------------------------
    # Parser 接口
    # ------------------------------------------------------------------
    def parse(
        self, artifact: RawArtifact, raw_bytes: bytes
    ) -> Tuple[List[RecordEnvelope], List[Dict]]:
        records: List[RecordEnvelope] = []
        rejects: List[Dict] = []
        parser_name = self._parser_name(artifact)
        for row, position, err in self._iter_items(artifact, raw_bytes):
            if err:
                rejects.append({
                    "artifact_id": artifact.artifact_id,
                    "reason": err,
                    "position": (position or {}).get("position", {}),
                })
                continue
            records.append(self._make_record(
                artifact, row, parser_name=parser_name, position=(position or {}).get("position", {}),
            ))
        return records, rejects

    def parse_stream(
        self,
        artifact: RawArtifact,
        raw_bytes: bytes,
        *,
        batch_size: int = 10_000,
    ) -> Iterator[Tuple[List[RecordEnvelope], List[Dict]]]:
        """逐批产出 (records, rejects)，支持大文件流式。"""
        batch: List[RecordEnvelope] = []
        rejects_batch: List[Dict] = []
        parser_name = self._parser_name(artifact)
        for row, position, err in self._iter_items(artifact, raw_bytes):
            if err:
                rejects_batch.append({
                    "artifact_id": artifact.artifact_id,
                    "reason": err,
                    "position": (position or {}).get("position", {}),
                })
                continue
            batch.append(self._make_record(
                artifact, row, parser_name=parser_name, position=(position or {}).get("position", {}),
            ))
            if len(batch) >= batch_size:
                yield batch, rejects_batch
                batch = []
                rejects_batch = []
        if batch or rejects_batch:
            yield batch, rejects_batch
