# -*- coding: utf-8 -*-
"""文档抽取的权威 JSONL、质量报告、血缘和 Manifest 交付。"""
from __future__ import annotations

import sys
import json
from collections import Counter
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Mapping, Sequence

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.worksheet import Worksheet

from src.data_prep.artifact_store import ArtifactStore
from src.data_prep.document_models import (
    ExtractedAggregate,
    ExtractedDocument,
    ExtractedField,
    ExtractedRecord,
    ExtractedTable,
    ExtractionSpec,
    ResultShape,
    ReviewTask,
)
from src.services.document_extraction import build_evidence_aggregate
from src.data_prep.models import (
    DatasetManifest,
    ManifestArtifactEntry,
    ManifestOutputEntry,
    OutputFormat,
    QualityDimensionResult,
    QualityReport,
    QualityResult,
    RawArtifact,
)


@dataclass(frozen=True)
class DocumentDelivery:
    manifest_path: str
    quality: QualityReport
    counts: dict[str, int]


def _relative(path: Path) -> str:
    return str(path).replace("\\", "/")


def _artifact_entry(
    store: ArtifactStore,
    rel_path: str | Path,
    kind: str,
) -> ManifestArtifactEntry:
    rel = _relative(Path(rel_path))
    path = store.resolve_path(rel)
    return ManifestArtifactEntry(
        artifact_id=f"{kind}:{path.stem}",
        kind=kind,
        path=rel,
        sha256=store.file_sha256(rel),
        size_bytes=path.stat().st_size,
    )


_HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
_STATUS_FILLS = {
    "found": PatternFill("solid", fgColor="E2F0D9"),
    "resolved": PatternFill("solid", fgColor="E2F0D9"),
    "not_found": PatternFill("solid", fgColor="F2F2F2"),
    "low_confidence": PatternFill("solid", fgColor="FFF2CC"),
    "pending": PatternFill("solid", fgColor="FFF2CC"),
    "conflict": PatternFill("solid", fgColor="FCE4D6"),
}


def _excel_value(value: Any) -> Any:
    if value is None:
        return ""
    if isinstance(value, (bool, int, float)):
        return value
    if isinstance(value, (dict, list, tuple, set)):
        value = json.dumps(value, ensure_ascii=False, default=str)
    text = str(value)
    if text.startswith(("=", "+", "-", "@")):
        return "'" + text
    return text


def _write_sheet(
    workbook: Workbook,
    name: str,
    headers: Sequence[str],
    rows: Sequence[Sequence[Any]],
) -> Worksheet:
    sheet = workbook.create_sheet(name)
    sheet.append(list(headers))
    for cell in sheet[1]:
        cell.fill = _HEADER_FILL
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center")
    for row in rows:
        sheet.append([_excel_value(value) for value in row])
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    status_columns = {
        index + 1
        for index, header in enumerate(headers)
        if header in {"status", "field_status", "overall"}
    }
    for row in sheet.iter_rows(min_row=2):
        for column in status_columns:
            cell = row[column - 1]
            fill = _STATUS_FILLS.get(str(cell.value or ""))
            if fill:
                cell.fill = fill
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    for index, header in enumerate(headers, start=1):
        values = [str(header)]
        values.extend(
            str(sheet.cell(row=row, column=index).value or "")
            for row in range(2, min(sheet.max_row, 100) + 1)
        )
        sheet.column_dimensions[sheet.cell(1, index).column_letter].width = min(
            50,
            max(12, max(len(value) for value in values) + 2),
        )
    return sheet


def _write_document_xlsx(
    store: ArtifactStore,
    task_id: str,
    *,
    spec: ExtractionSpec,
    raw_artifacts: Sequence[RawArtifact],
    field_rows: Sequence[Mapping[str, Any]],
    evidence_rows: Sequence[Mapping[str, Any]],
    review_rows: Sequence[Mapping[str, Any]],
    quality: QualityReport,
    parse_rejects: Sequence[Mapping[str, Any]],
    record_rows: Sequence[Mapping[str, Any]] = (),
    tables: Sequence[ExtractedTable] = (),
    documents: Sequence[ExtractedDocument] = (),
    aggregates: Sequence[ExtractedAggregate] = (),
) -> str:
    """生成非权威的业务查看副本；权威记录仍是 JSONL。"""
    workbook = Workbook()
    workbook.remove(workbook.active)
    _write_sheet(
        workbook,
        "Fields",
        (
            "field_name",
            "status",
            "value",
            "review_reason",
            "evidence_count",
            "artifact_id",
            "page",
            "quote",
            "confidence",
        ),
        [
            (
                field.get("name"),
                field.get("status"),
                field.get("value"),
                field.get("review_reason"),
                len(field.get("evidence_refs") or []),
                (field.get("evidence_refs") or [{}])[0].get("artifact_id"),
                (field.get("evidence_refs") or [{}])[0].get("page"),
                (field.get("evidence_refs") or [{}])[0].get("quote"),
                (field.get("evidence_refs") or [{}])[0].get("confidence"),
            )
            for field in field_rows
        ],
    )
    if record_rows:
        record_headers = [
            "_record_id",
            "_status",
            "_source_artifact_ids",
            *[item.name for item in spec.fields],
        ]
        _write_sheet(
            workbook,
            "Records",
            record_headers,
            [
                (
                    row.get("record_id"),
                    row.get("status"),
                    row.get("source_artifact_ids"),
                    *[
                        (row.get("values") or {}).get(field.name)
                        for field in spec.fields
                    ],
                )
                for row in record_rows
            ],
        )
    if documents:
        document_rows = []
        for document in documents:
            chunks = [
                document.content[index:index + 30_000]
                for index in range(0, len(document.content), 30_000)
            ]
            document_rows.extend(
                (
                    document.document_id,
                    document.title,
                    chunk_no,
                    document.source_artifact_ids,
                    chunk,
                )
                for chunk_no, chunk in enumerate(chunks, start=1)
            )
        _write_sheet(
            workbook,
            "Documents",
            ("document_id", "title", "chunk_no", "source_artifact_ids", "content"),
            document_rows,
        )
    if aggregates:
        _write_sheet(
            workbook,
            "Aggregate",
            ("aggregate_id", "field_name", "status", "value", "evidence_count"),
            [
                (
                    aggregate.aggregate_id,
                    field.name,
                    field.status.value,
                    field.value,
                    len(field.evidence_refs),
                )
                for aggregate in aggregates
                for field in aggregate.fields
            ],
        )
    used_sheet_names = set(workbook.sheetnames)
    for index, table in enumerate(tables, start=1):
        base = (table.name or f"Table{index}")[:25]
        name = base
        suffix = 1
        while name in used_sheet_names:
            suffix += 1
            name = f"{base[:21]}-{suffix}"
        used_sheet_names.add(name)
        _write_sheet(
            workbook,
            name,
            table.columns,
            [
                tuple(row.get(column) for column in table.columns)
                for row in table.rows
            ],
        )
    _write_sheet(
        workbook,
        "Evidence",
        (
            "field_name",
            "field_status",
            "field_value",
            "artifact_id",
            "element_id",
            "page",
            "bbox",
            "quote",
            "quote_sha256",
            "extractor",
            "extractor_version",
            "confidence",
            "raw_result_ref",
        ),
        [
            tuple(row.get(header) for header in (
                "field_name",
                "field_status",
                "field_value",
                "artifact_id",
                "element_id",
                "page",
                "bbox",
                "quote",
                "quote_sha256",
                "extractor",
                "extractor_version",
                "confidence",
                "raw_result_ref",
            ))
            for row in evidence_rows
        ],
    )
    _write_sheet(
        workbook,
        "Review",
        (
            "review_task_id",
            "artifact_id",
            "page",
            "field_name",
            "status",
            "reasons",
            "candidate_count",
            "resolution",
        ),
        [
            (
                row.get("task_id"),
                row.get("artifact_id"),
                row.get("page"),
                row.get("field_name"),
                row.get("status"),
                row.get("reasons"),
                len(row.get("candidates") or []),
                row.get("resolution"),
            )
            for row in review_rows
        ],
    )
    _write_sheet(
        workbook,
        "Quality",
        ("category", "name", "value", "threshold", "passed", "detail"),
        [
            (
                "summary",
                "overall",
                quality.overall.value,
                "",
                quality.overall.value != "fail",
                "",
            ),
            *[
                (
                    "dimension",
                    dimension.name,
                    dimension.value,
                    dimension.threshold,
                    dimension.passed,
                    "",
                )
                for dimension in quality.dimensions
            ],
            *[
                ("issue", "issue", "", "", False, issue)
                for issue in quality.issues
            ],
        ],
    )
    manifest_rows = [
        ("task_id", task_id),
        ("spec_version", spec.spec_version),
        ("engine", "document_extraction_v1"),
        ("authoritative_format", "JSONL"),
        *[
            (f"count.{key}", value)
            for key, value in sorted(quality.counts.items())
        ],
        *[
            (f"raw_artifact.{index}", artifact.artifact_id)
            for index, artifact in enumerate(raw_artifacts, start=1)
        ],
    ]
    _write_sheet(workbook, "Manifest", ("key", "value"), manifest_rows)
    reject_headers = (
        "artifact_id",
        "reason",
        "position",
        "page_kind",
        "document_kind",
        "ocr_error",
    )
    _write_sheet(
        workbook,
        "Rejects",
        reject_headers,
        [
            tuple(row.get(header) for header in reject_headers)
            for row in parse_rejects
        ],
    )
    rel_path = f"{task_id}/extraction/document_extraction.xlsx"
    target = store.resolve_path(rel_path)
    target.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(target)
    return rel_path


def write_document_delivery(
    store: ArtifactStore,
    task_id: str,
    *,
    spec: ExtractionSpec,
    raw_artifacts: Sequence[RawArtifact],
    fields: Sequence[ExtractedField],
    review_tasks: Sequence[ReviewTask],
    parse_rejects: Sequence[Mapping[str, Any]] = (),
    records: Sequence[ExtractedRecord] = (),
    tables: Sequence[ExtractedTable] = (),
    documents: Sequence[ExtractedDocument] = (),
    coverage: Mapping[str, int] | None = None,
    raw_tables: Sequence[ExtractedTable] | None = None,
    table_recipe_audit: Mapping[str, Any] | None = None,
) -> DocumentDelivery:
    """重建文档任务交付；人工裁决后可安全重复调用。"""
    field_rows = [item.model_dump(mode="json") for item in fields]
    record_rows = [
        {
            **item.model_dump(mode="json"),
            "values": item.values,
        }
        for item in records
    ]
    record_fields = [
        field
        for record in records
        for field in record.fields
    ]
    aggregates = (
        [build_evidence_aggregate(spec, fields)]
        if spec.result_contract.shape == ResultShape.AGGREGATE and fields
        else []
    )
    all_fields = [*fields, *record_fields]
    review_rows = [item.model_dump(mode="json") for item in review_tasks]
    evidence_rows = [
        {
            "field_name": field.name,
            "field_status": field.status.value,
            "field_value": field.value,
            **ref.model_dump(mode="json"),
        }
        for field in all_fields
        for ref in field.evidence_refs
    ]
    evidence_rows.extend(
        {
            "field_name": f"document:{document.document_id}",
            "field_status": "found",
            "field_value": document.title,
            **ref.model_dump(mode="json"),
        }
        for document in documents
        for ref in document.evidence_refs
    )
    lineage_rows = [
        {
            "field_name": field.name,
            "status": field.status.value,
            "artifact_ids": sorted({
                ref.artifact_id for ref in field.evidence_refs
            }),
            "element_ids": [ref.element_id for ref in field.evidence_refs],
        }
        for field in all_fields
    ]
    lineage_rows.extend(
        {
            "field_name": f"document:{document.document_id}",
            "status": "found",
            "artifact_ids": document.source_artifact_ids,
            "element_ids": [ref.element_id for ref in document.evidence_refs],
        }
        for document in documents
    )

    fields_json_path = store.write_json(
        task_id, "extraction/extracted_fields.json", field_rows
    )
    reviews_path = store.write_json(
        task_id, "extraction/review_tasks.json", review_rows
    )
    spec_path = store.write_json(
        task_id, "extraction/extraction_spec.json", spec.model_dump(mode="json")
    )
    fields_path = store.write_jsonl(
        task_id, "extraction", "extracted_fields.jsonl", field_rows
    )
    records_path = store.write_jsonl(
        task_id, "extraction", "extracted_records.jsonl", record_rows
    )
    tables_path = store.write_json(
        task_id,
        "extraction/extracted_tables.json",
        [item.model_dump(mode="json") for item in tables],
    )
    documents_path = store.write_json(
        task_id,
        "extraction/extracted_documents.json",
        [item.model_dump(mode="json") for item in documents],
    )
    aggregates_path = store.write_json(
        task_id,
        "extraction/extracted_aggregates.json",
        [
            {
                **item.model_dump(mode="json"),
                "values": item.values,
            }
            for item in aggregates
        ],
    )
    raw_tables_path = None
    if raw_tables is not None:
        raw_tables_path = store.write_json(
            task_id,
            "extraction/extracted_tables_raw.json",
            [item.model_dump(mode="json") for item in raw_tables],
        )
    table_recipe_path = None
    if table_recipe_audit is not None:
        table_recipe_path = store.write_json(
            task_id,
            "extraction/table_recipe_audit.json",
            dict(table_recipe_audit),
        )
    evidence_path = store.write_jsonl(
        task_id, "extraction", "evidence.jsonl", evidence_rows
    )
    lineage_path = store.write_jsonl(
        task_id, "extraction", "lineage.jsonl", lineage_rows
    )
    rejects_path = None
    if parse_rejects:
        rejects_path = store.write_json(
            task_id, "extraction/parse_rejects.json", list(parse_rejects)
        )

    schema_path = store.write_schema(task_id, {
        "spec_version": spec.spec_version,
        "fields": [item.model_dump(mode="json") for item in spec.fields],
        "record_count": (
            len(records)
            or sum(len(item.rows) for item in tables)
            or len(documents)
            or len(aggregates)
            or len(fields)
        ),
        "inferred": False,
    })
    status_counts = Counter(field.status.value for field in all_fields)
    field_found = status_counts["found"]
    found = field_found + len(documents)
    evidence_complete = sum(
        field.status.value == "found" and bool(field.evidence_refs)
        for field in all_fields
    ) + sum(bool(document.evidence_refs) for document in documents)
    allowed_artifacts = set(spec.discovery.artifact_ids)
    cross_document_refs = sum(
        ref.artifact_id not in allowed_artifacts
        for field in all_fields
        for ref in field.evidence_refs
    ) + sum(
        ref.artifact_id not in allowed_artifacts
        for document in documents
        for ref in document.evidence_refs
    )
    required_names = (
        set()
        if spec.result_contract.shape == ResultShape.TABLES
        else {item.name for item in spec.fields if item.required}
    )
    required_complete = sum(
        field.name in required_names and field.status.value == "found"
        for field in all_fields
    )
    required_slots = len(required_names) * (len(records) if records else 1)
    resolved_reviews = sum(item.status == "resolved" for item in review_tasks)
    pending_reviews = len(review_tasks) - resolved_reviews
    unsafe_found = found - evidence_complete
    unresolved_field_states = status_counts["conflict"] + status_counts["low_confidence"]
    table_row_count = sum(len(item.rows) for item in tables)
    expected_output_empty = (
        spec.result_contract.shape == ResultShape.TABLES
        and table_row_count == 0
    ) or (
        spec.result_contract.shape == ResultShape.RECORDS
        and not records
    ) or (
        spec.result_contract.shape == ResultShape.DOCUMENT
        and not documents
    ) or (
        spec.result_contract.shape == ResultShape.AGGREGATE
        and not any(
            field.status.value == "found"
            for aggregate in aggregates
            for field in aggregate.fields
        )
    )
    if unsafe_found or cross_document_refs or expected_output_empty:
        overall = QualityResult.FAIL
    elif (
        pending_reviews
        or required_complete < required_slots
        or unresolved_field_states
        or parse_rejects
    ):
        overall = QualityResult.WARN
    else:
        overall = QualityResult.PASS
    counts = {
        "extracted_fields": len(fields),
        "extracted_records": len(records),
        "extracted_tables": len(tables),
        "extracted_table_rows": table_row_count,
        "extracted_documents": len(documents),
        "extracted_aggregates": len(aggregates),
        "found": found,
        "not_found": status_counts["not_found"],
        "conflict": status_counts["conflict"],
        "low_confidence": status_counts["low_confidence"],
        "evidence_refs": len(evidence_rows),
        "review_tasks": len(review_tasks),
        "review_tasks_pending": pending_reviews,
        "parse_rejects": len(parse_rejects),
        **dict(coverage or {}),
    }
    issues = []
    if unsafe_found:
        issues.append(f"{unsafe_found} 个 found 字段缺少证据")
    if cross_document_refs:
        issues.append(f"{cross_document_refs} 条证据越过任务文档范围")
    if pending_reviews:
        issues.append(f"{pending_reviews} 项人工复核尚未完成")
    if required_complete < required_slots:
        issues.append("必填字段未全部安全找到")
    if unresolved_field_states:
        issues.append(f"{unresolved_field_states} 个字段仍为冲突或低置信度")
    if parse_rejects:
        issues.append(f"{len(parse_rejects)} 个解析异常")
    if expected_output_empty:
        expected_name = (
            "表格行"
            if spec.result_contract.shape == ResultShape.TABLES
            else "记录"
            if spec.result_contract.shape == ResultShape.RECORDS
            else "连续文档"
            if spec.result_contract.shape == ResultShape.DOCUMENT
            else "有效汇总指标"
        )
        issues.append(f"任务要求全量输出，但未产出任何{expected_name}")
    quality = QualityReport(
        task_id=task_id,
        overall=overall,
        dimensions=[
            QualityDimensionResult(
                name="证据完整率",
                value=evidence_complete / found if found else 1.0,
                threshold=1.0,
                passed=unsafe_found == 0,
            ),
            QualityDimensionResult(
                name="必填字段完整率",
                value=required_complete / required_slots if required_slots else 1.0,
                threshold=1.0,
                passed=required_complete == required_slots,
            ),
            QualityDimensionResult(
                name="复核完成率",
                value=resolved_reviews / len(review_tasks) if review_tasks else 1.0,
                threshold=1.0,
                passed=pending_reviews == 0,
            ),
            QualityDimensionResult(
                name="跨文档隔离",
                value=1.0 if cross_document_refs == 0 else 0.0,
                threshold=1.0,
                passed=cross_document_refs == 0,
            ),
            QualityDimensionResult(
                name="有效结果",
                value=0.0 if expected_output_empty else 1.0,
                threshold=1.0,
                passed=not expected_output_empty,
            ),
        ],
        issues=issues,
        counts=counts,
    )
    quality_path = store.write_quality(task_id, quality)
    xlsx_path = None
    if quality.overall != QualityResult.FAIL:
        xlsx_path = _write_document_xlsx(
            store,
            task_id,
            spec=spec,
            raw_artifacts=raw_artifacts,
            field_rows=field_rows,
            evidence_rows=evidence_rows,
            review_rows=review_rows,
            quality=quality,
            parse_rejects=parse_rejects,
            record_rows=record_rows,
            tables=tables,
            documents=documents,
            aggregates=aggregates,
        )

    generated_artifacts = [
        _artifact_entry(store, fields_json_path, "extracted_fields_json"),
        _artifact_entry(store, records_path, "extracted_records"),
        _artifact_entry(store, tables_path, "extracted_tables"),
        _artifact_entry(store, documents_path, "extracted_documents"),
        _artifact_entry(store, aggregates_path, "extracted_aggregates"),
        _artifact_entry(store, evidence_path, "evidence"),
        _artifact_entry(store, reviews_path, "review_tasks"),
        _artifact_entry(store, spec_path, "extraction_spec"),
        _artifact_entry(store, schema_path, "schema"),
        _artifact_entry(store, quality_path, "quality"),
        _artifact_entry(store, lineage_path, "lineage"),
    ]
    if raw_tables_path is not None:
        generated_artifacts.append(
            _artifact_entry(store, raw_tables_path, "extracted_tables_raw")
        )
    if table_recipe_path is not None:
        generated_artifacts.append(
            _artifact_entry(store, table_recipe_path, "table_recipe_audit")
        )
    if xlsx_path is not None:
        generated_artifacts.append(
            _artifact_entry(store, xlsx_path, "document_xlsx")
        )
    decisions_path = store.task_dir(task_id) / "extraction/review_decisions.jsonl"
    if decisions_path.exists():
        generated_artifacts.append(_artifact_entry(
            store,
            decisions_path.relative_to(store.root),
            "review_decisions",
        ))
    if rejects_path is not None:
        generated_artifacts.append(_artifact_entry(store, rejects_path, "rejects"))

    authoritative_source = {
        ResultShape.RECORDS: records_path,
        ResultShape.TABLES: tables_path,
        ResultShape.DOCUMENT: documents_path,
        ResultShape.AGGREGATE: aggregates_path,
    }.get(spec.result_contract.shape, fields_path)
    authoritative_path = _relative(authoritative_source)
    authoritative_format = (
        OutputFormat.JSONL
        if spec.result_contract.shape in {ResultShape.FIELDS, ResultShape.RECORDS}
        else OutputFormat.JSON
    )
    authoritative_records = (
        len(records)
        or sum(len(item.rows) for item in tables)
        or len(documents)
        or len(aggregates)
        or len(fields)
    )
    manifest = DatasetManifest(
        task_id=task_id,
        spec_version=spec.spec_version,
        artifacts=[
            *[
                ManifestArtifactEntry(
                    artifact_id=item.artifact_id,
                    kind="raw",
                    path=item.storage_path,
                    sha256=item.sha256,
                    size_bytes=item.size_bytes,
                )
                for item in raw_artifacts
            ],
            *generated_artifacts,
        ],
        outputs=[] if quality.overall == QualityResult.FAIL else [
            ManifestOutputEntry(
                format=authoritative_format,
                path=authoritative_path,
                sha256=store.file_sha256(authoritative_path),
                records=authoritative_records,
            ),
            ManifestOutputEntry(
                format=OutputFormat.XLSX,
                path=xlsx_path,
                sha256=store.file_sha256(xlsx_path),
                records=(
                    len(records)
                    or sum(len(item.rows) for item in tables)
                    or len(documents)
                    or len(aggregates)
                    or len(fields)
                ),
            ),
        ],
        record_counts=counts,
        schema_ref=_relative(schema_path),
        quality_ref=_relative(quality_path),
        lineage_ref=lineage_path,
        environment={
            "python": (
                f"{sys.version_info.major}.{sys.version_info.minor}."
                f"{sys.version_info.micro}"
            ),
            "engine": "document_extraction_v1",
            "created_at": datetime.now(timezone.utc).isoformat(),
        },
    )
    manifest_path = store.write_json(
        task_id, "manifest.json", manifest.model_dump(mode="json")
    )
    return DocumentDelivery(
        manifest_path=_relative(manifest_path),
        quality=quality,
        counts=counts,
    )
