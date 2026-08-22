# -*- coding: utf-8 -*-
"""Phase 4B 批次 7：正式工作台后台运行、隔离和交付门禁。"""
from __future__ import annotations

import asyncio
import csv
from contextlib import asynccontextmanager
from datetime import datetime, timedelta
import io
import json
from pathlib import Path
import threading
import time
import zipfile

from docx import Document
from fastapi import FastAPI
from fastapi.testclient import TestClient
from openpyxl import load_workbook
from pypdf import PdfReader
import pytest

import src.api.auth as auth_mod
import src.api.semantic_workspace_runtime as runtime_mod
from src.semantic_harness.delivery import service as delivery_service
from src.api.auth import get_current_user, get_store
from src.api.routes import semantic_plans, semantic_workspace
from src.api.semantic_workspace_runtime import SemanticWorkspaceManager
from src.api.store import WebUIStore
from src.config.settings import settings
from src.semantic_harness.compiler_models import (
    CompileRequest,
    PlanSemanticsDraft,
)
from src.semantic_harness.models import (
    Ambiguity,
    CombineMode,
    CombineSpec,
    ContentPolicy,
    DeliveryFormat,
    DeliverySpec,
    OperationSpec,
    OperationType,
    PostconditionSpec,
    PredicateOperator,
    PredicatePostcondition,
    ProjectionField,
    TaskFamily,
)
from src.services.upload_store import UploadStore
from tests.test_semantic_plan_api import ApiFakeGenerator


def test_workspace_revision_freezes_table_output_contract(tmp_path: Path) -> None:
    store = WebUIStore(str(tmp_path / "workspace-contract.db"))
    contract = [{
        "format": "json",
        "exact_columns": ["name", "amount"],
        "json_shape": "records",
    }]

    task = store.create_semantic_workspace_task(
        "user-a",
        task_id="workspace-contract",
        title="结构化交付",
        objective_text="输出姓名和金额",
        upload_ids=[],
        output_formats=["json"],
        provider="local",
        model="local-model",
        external_api_confirmed=False,
        table_output_contracts=contract,
    )
    revision = store.get_semantic_workspace_revision(
        "user-a",
        "workspace-contract",
        1,
    )
    revised = store.create_semantic_workspace_revision(
        "user-a",
        "workspace-contract",
        objective_text="仍然输出姓名和金额",
        output_formats=["json"],
        change_summary="只调整筛选范围",
    )

    assert task["table_output_contracts"] == contract
    assert revision is not None
    assert revision["table_output_contracts"] == contract
    assert revised["table_output_contracts"] == contract


def test_concurrent_revision_uses_transactional_expected_revision(
    tmp_path: Path,
) -> None:
    store = WebUIStore(str(tmp_path / "workspace-revision-race.db"))
    store.create_semantic_workspace_task(
        "user-a",
        task_id="workspace-race",
        title="并发版本",
        objective_text="初始目标",
        upload_ids=[],
        output_formats=["json"],
        provider="local",
        model="local-model",
        external_api_confirmed=False,
    )
    barrier = threading.Barrier(2)
    outcomes: list[str] = []

    def create_revision(instruction: str) -> None:
        barrier.wait()
        try:
            store.create_semantic_workspace_revision(
                "user-a",
                "workspace-race",
                objective_text=instruction,
                output_formats=["json"],
                change_summary=instruction,
                expected_revision=2,
            )
        except RuntimeError:
            outcomes.append("conflict")
        else:
            outcomes.append("created")

    threads = [
        threading.Thread(target=create_revision, args=(f"修改-{index}",))
        for index in range(2)
    ]
    for thread in threads:
        thread.start()
    for thread in threads:
        thread.join(timeout=5)

    assert sorted(outcomes) == ["conflict", "created"]
    task = store.get_semantic_workspace_task("user-a", "workspace-race")
    assert task is not None
    assert task["active_revision"] == 2
    assert store.get_semantic_workspace_revision(
        "user-a", "workspace-race", 3
    ) is None


def _client(tmp_path, monkeypatch, *, generator=None):
    monkeypatch.setattr(
        settings, "webui_db_path", str(tmp_path / "workspace.db")
    )
    monkeypatch.setattr(
        settings, "data_prep_upload_root", str(tmp_path / "uploads")
    )
    monkeypatch.setattr(
        settings,
        "semantic_execution_root",
        str(tmp_path / "executions"),
    )
    auth_mod._store = None
    monkeypatch.setattr(
        semantic_plans,
        "_build_generator",
        lambda **_: generator or ApiFakeGenerator(),
    )
    manager = SemanticWorkspaceManager()
    monkeypatch.setattr(runtime_mod, "_manager", manager)
    current_user = {"value": "user-a"}

    @asynccontextmanager
    async def lifespan(_app):
        manager.start()
        try:
            yield
        finally:
            await manager.stop()

    app = FastAPI(lifespan=lifespan)
    app.include_router(semantic_workspace.router)
    app.dependency_overrides[get_current_user] = lambda: {
        "user_id": current_user["value"]
    }
    return TestClient(app), current_user


@pytest.mark.parametrize(
    ("delivery_format", "extension", "expected_kind"),
    [
        (DeliveryFormat.JSON, "json", "document"),
        (DeliveryFormat.JSONL, "jsonl", "table"),
        (DeliveryFormat.CSV, "csv", "table"),
        (DeliveryFormat.XLSX, "xlsx", "table"),
        (DeliveryFormat.PARQUET, "parquet", "table"),
        (DeliveryFormat.DOCX, "docx", "document"),
        (DeliveryFormat.PDF, "pdf", "document"),
        (DeliveryFormat.HTML, "html", "document"),
        (DeliveryFormat.MARKDOWN, "md", "document"),
        (DeliveryFormat.TXT, "txt", "document"),
        (DeliveryFormat.PPTX, "pptx", "document"),
    ],
)
def test_every_selectable_delivery_format_has_online_preview(
    tmp_path: Path,
    delivery_format: DeliveryFormat,
    extension: str,
    expected_kind: str,
) -> None:
    content = delivery_service.CanonicalContent(
        title="费用结果",
        columns=("姓名", "费用合计"),
        rows=({"姓名": "董琳", "费用合计": 200},),
        sections=(("董琳", "费用合计 200"),),
        raw={"董琳": {"费用合计": 200}},
    )
    path = tmp_path / f"result.{extension}"
    delivery_service._RENDERERS[delivery_format](path, content)

    preview = semantic_workspace._preview_result_file(
        path,
        lineage_path=None,
        offset=0,
        limit=10,
        search="",
        sort_by=None,
        sort_direction="asc",
    )

    assert preview["kind"] == expected_kind
    assert preview["total"] >= 1


def _upload(tmp_path):
    rows = ["姓名,核销工作量天数,工作量费用"]
    rows.extend(
        f"谢超群,{index / 2:.1f},{index * 100:.2f}"
        for index in range(1, 12)
    )
    return UploadStore(
        root=str(tmp_path / "uploads"),
        max_bytes=10 * 1024 * 1024,
    ).save_bytes(
        "user-a",
        "workload.csv",
        ("\n".join(rows) + "\n").encode("utf-8"),
        media_type="text/csv",
    )


def _upload_contract(tmp_path):
    fixture = (
        Path(__file__).parent
        / "fixtures"
        / "semantic_harness"
        / "public"
        / "batch0"
        / "documents"
        / "contract.docx"
    )
    return UploadStore(
        root=str(tmp_path / "uploads"),
        max_bytes=10 * 1024 * 1024,
    ).save_bytes(
        "user-a",
        "contract.docx",
        fixture.read_bytes(),
        media_type=(
            "application/vnd.openxmlformats-officedocument."
            "wordprocessingml.document"
        ),
    )


class ScopedDocumentClarificationGenerator:
    """第二轮故意漏掉范围，验证编译器必须从上一版保留。"""

    provider = "local"
    model = "scoped-document-fixture"
    prompt_version = "stp-v1-test"
    prompt_sha256 = "8" * 64

    async def generate(
        self,
        request: CompileRequest,
        *,
        diagnostics,
        attempt: int,
    ) -> PlanSemanticsDraft:
        del diagnostics, attempt
        if request.clarification is None:
            return PlanSemanticsDraft(
                task_family=TaskFamily.EXTRACT,
                normalized_objective="提取付款条款并输出文本",
                section_patterns=("付款条款",),
                record_grain="条款",
                content_policy=ContentPolicy.VERBATIM,
                delivery=DeliverySpec(formats=(DeliveryFormat.TXT,)),
                ambiguities=(
                    Ambiguity(
                        ambiguity_id="extract.mode",
                        question="需要逐字原文还是摘要？",
                        candidates=("逐字原文", "摘要"),
                    ),
                ),
            )
        return PlanSemanticsDraft(
            task_family=TaskFamily.EXTRACT,
            normalized_objective="逐字提取条款原文",
            content_policy=ContentPolicy.VERBATIM,
            delivery=DeliverySpec(formats=(DeliveryFormat.DOCX,)),
        )


class TruncatedThenInvalidPlanGenerator:
    """复现真实任务：两次输出截断，最后一次计划仍未通过校验。"""

    provider = "local"
    model = "truncated-plan-fixture"
    prompt_version = "stp-v1-test"
    prompt_sha256 = "9" * 64

    def __init__(self) -> None:
        self.calls = 0

    async def generate(
        self,
        request: CompileRequest,
        *,
        diagnostics,
        attempt: int,
    ) -> PlanSemanticsDraft:
        del request, diagnostics, attempt
        self.calls += 1
        if self.calls <= 2:
            raise RuntimeError(
                "IncompleteOutputException: The output is incomplete "
                "due to a max_tokens length limit."
            )
        return PlanSemanticsDraft(
            task_family=TaskFamily.EXTRACT,
            normalized_objective="提取并汇总商务条款",
            content_policy=ContentPolicy.VERBATIM,
            delivery=DeliverySpec(formats=(DeliveryFormat.TXT,)),
        )


class Batch8AFileScenarioGenerator:
    """四条公开文件闭环的确定性语义生成器。"""

    provider = "local"
    model = "batch8a-file-scenarios"
    prompt_version = "stp-v1-batch8a"
    prompt_sha256 = "8" * 64

    async def generate(
        self,
        request: CompileRequest,
        *,
        diagnostics,
        attempt: int,
    ) -> PlanSemanticsDraft:
        del diagnostics, attempt
        objective = request.objective_text
        if "商务条款" in objective:
            return PlanSemanticsDraft(
                task_family=TaskFamily.EXTRACT,
                normalized_objective="提取并汇总商务条款",
                accepted_formats=("docx",),
                selection=({
                    "field": "正文",
                    "operator": PredicateOperator.CONTAINS,
                    "value": "商务条款",
                },),
                content_policy=ContentPolicy.VERBATIM,
                delivery=DeliverySpec(
                    formats=(DeliveryFormat.TXT,),
                    output_name="商务条款汇总",
                ),
                postconditions=PostconditionSpec(
                    predicates=(
                        PredicatePostcondition(
                            field="正文",
                            operator=PredicateOperator.CONTAINS,
                            value="商务条款",
                        ),
                    ),
                ),
            )
        if "核查违约责任" in objective:
            return PlanSemanticsDraft(
                task_family=TaskFamily.AUDIT,
                normalized_objective="核查违约责任并保留证据",
                whole_document=True,
                accepted_formats=("pdf",),
                operations=(
                    OperationSpec(
                        operation=OperationType.AUDIT,
                        params={
                            "rules": [{
                                "label": "违约责任",
                                "query": "违约责任",
                                "operator": "contains",
                                "value": "千分之一",
                            }],
                        },
                    ),
                ),
                content_policy=ContentPolicy.ANALYZED,
                delivery=DeliverySpec(
                    formats=(
                        DeliveryFormat.DOCX,
                        DeliveryFormat.PDF,
                    ),
                    output_name="违约责任核查",
                ),
            )
        if "筛选排序" in objective:
            return PlanSemanticsDraft(
                task_family=TaskFamily.TABULAR_TRANSFORM,
                normalized_objective="筛选示例人员甲并按费用降序",
                table_scope="all_detected_tables",
                accepted_formats=("xlsx",),
                selection=({
                    "field": "姓名",
                    "operator": PredicateOperator.EQ,
                    "value": "示例人员甲",
                },),
                projection=(
                    ProjectionField(name="姓名"),
                    ProjectionField(name="工作量费用"),
                ),
                record_grain="source_detail_row",
                operations=(
                    OperationSpec(
                        operation=OperationType.SORT,
                        params={
                            "columns": ["工作量费用"],
                            "direction": "desc",
                        },
                    ),
                ),
                combine=CombineSpec(mode=CombineMode.ONE_TABLE),
                content_policy=ContentPolicy.VERBATIM,
                delivery=DeliverySpec(
                    formats=(DeliveryFormat.XLSX,),
                    output_name="人员费用明细",
                ),
                postconditions=PostconditionSpec(
                    table_count=1,
                    exact_visible_columns=("姓名", "工作量费用"),
                    predicates=(
                        PredicatePostcondition(
                            field="姓名",
                            operator=PredicateOperator.EQ,
                            value="示例人员甲",
                        ),
                    ),
                ),
            )
        return PlanSemanticsDraft(
            task_family=TaskFamily.TABULAR_TRANSFORM,
            normalized_objective="按人员去重并汇总费用",
            table_scope="all_detected_tables",
            accepted_formats=("csv",),
            projection=(
                ProjectionField(name="姓名"),
                ProjectionField(name="费用合计"),
            ),
            record_grain="每名人员一行",
            operations=(
                OperationSpec(
                    operation=OperationType.DEDUPLICATE,
                    params={"keys": ["姓名", "项目", "月份"]},
                ),
                OperationSpec(
                    operation=OperationType.GROUP,
                    params={"columns": ["姓名"]},
                ),
                OperationSpec(
                    operation=OperationType.AGGREGATE,
                    params={
                        "aggregates": [{
                            "function": "sum",
                            "column": "工作量费用",
                            "output": "费用合计",
                        }],
                    },
                ),
            ),
            combine=CombineSpec(mode=CombineMode.ONE_TABLE),
            content_policy=ContentPolicy.NORMALIZED,
            delivery=DeliverySpec(
                formats=(DeliveryFormat.CSV, DeliveryFormat.JSON),
                output_name="人员费用汇总",
            ),
            postconditions=PostconditionSpec(
                table_count=1,
                exact_visible_columns=("姓名", "费用合计"),
            ),
        )


def _wait(client: TestClient, task_id: str, expected: set[str]):
    deadline = time.monotonic() + 30
    while time.monotonic() < deadline:
        response = client.get(
            f"/api/semantic-workspace/tasks/{task_id}"
        )
        assert response.status_code == 200, response.text
        payload = response.json()
        if (
            payload["status"] in expected
            and not (
                payload["status"] == "needs_input"
                and payload.get("question") is None
            )
        ):
            return payload
        time.sleep(0.1)
    raise AssertionError(f"任务未进入期望状态：{expected}")


def test_workspace_explains_bounded_compile_failure_to_task_owner(
    tmp_path,
    monkeypatch,
):
    generator = TruncatedThenInvalidPlanGenerator()
    client, _ = _client(tmp_path, monkeypatch, generator=generator)
    upload = _upload_contract(tmp_path)

    with client:
        created = client.post(
            "/api/semantic-workspace/tasks",
            json={
                "objective_text": (
                    "这个 Word 里的商务条款有哪些？请汇总并输出 TXT"
                ),
                "upload_ids": [upload.upload_id],
                "output_formats": ["txt"],
                "provider": "local",
                "model": generator.model,
            },
        )
        assert created.status_code == 202, created.text

        task = _wait(client, created.json()["task_id"], {"failed"})

    assert generator.calls == 3
    assert task["question"] is None
    assert task["failure"] is not None, json.dumps(
        task,
        ensure_ascii=False,
        indent=2,
    )
    assert task["failure"] == {
        "error_code": "STP_COMPILE_FAILED",
        "stage": "interpret",
        "cause_summary": "本地模型两次输出被截断，最后生成的计划未通过校验",
        "attempt_count": 3,
        "elapsed_ms": task["failure"]["elapsed_ms"],
        "source_read": False,
        "intermediate_created": False,
        "delivery_published": False,
        "next_actions": ["修改要求后重试", "检查本地模型配置"],
        "diagnostic_ref": task["plan_id"],
    }
    assert task["failure"]["elapsed_ms"] >= 0
    assert any(
        event["stage"] == "interpret"
        and event["event_type"] == "stage_failed"
        and "两次输出被截断" in event["summary"]
        for event in task["events"]
    )
    assert all(
        not (
            event["stage"] == "interpret"
            and event["event_type"] == "stage_completed"
            and "需要一项补充" in event["summary"]
        )
        for event in task["events"]
    )


def test_workspace_runs_in_background_previews_and_bundles(
    tmp_path,
    monkeypatch,
):
    client, _ = _client(tmp_path, monkeypatch)
    upload = _upload(tmp_path)
    with client:
        started = time.monotonic()
        created = client.post(
            "/api/semantic-workspace/tasks",
            json={
                "objective_text": (
                    "只提取谢超群的数据，只保留核销工作量天数和工作量费用"
                ),
                "upload_ids": [upload.upload_id],
                "output_formats": ["xlsx"],
            },
        )
        assert created.status_code == 202, created.text
        assert time.monotonic() - started < 2
        task_id = created.json()["task_id"]
        completed = _wait(client, task_id, {"completed"})
        assert completed["delivery"]["outputs"]
        assert completed["summary"].startswith("目标：")
        assert any(
            event["event_type"] == "task_completed"
            for event in completed["events"]
        )

        preview = client.get(
            f"/api/semantic-workspace/tasks/{task_id}/preview"
        )
        assert preview.status_code == 200, preview.text
        assert preview.json()["kind"] == "table"
        assert preview.json()["total"] == 11
        assert preview.json()["rows"][0]["__lineage"]

        bundle = client.get(
            f"/api/semantic-workspace/tasks/{task_id}/bundle"
        )
        assert bundle.status_code == 200, bundle.text
        with zipfile.ZipFile(io.BytesIO(bundle.content)) as archive:
            names = set(archive.namelist())
        assert {"manifest.json", "qa.json", "trace.json"} <= names
        assert not any(name.startswith("sources/") for name in names)


def test_workspace_rejects_mixed_document_and_table_inputs(
    tmp_path,
    monkeypatch,
):
    client, _ = _client(tmp_path, monkeypatch)
    document = _upload_contract(tmp_path)
    table = _upload(tmp_path)

    with client:
        response = client.post(
            "/api/semantic-workspace/tasks",
            json={
                "objective_text": "对照合同条款和工作量表生成汇总",
                "upload_ids": [document.upload_id, table.upload_id],
                "output_formats": ["xlsx"],
            },
        )

        assert response.status_code == 422
        assert "文档和表格请分别创建任务" in response.json()["detail"]
        assert client.get("/api/semantic-workspace/tasks").json() == []


def test_workspace_rejects_prompt_and_selector_output_conflict(
    tmp_path,
    monkeypatch,
):
    client, _ = _client(tmp_path, monkeypatch)
    upload = _upload(tmp_path)

    with client:
        response = client.post(
            "/api/semantic-workspace/tasks",
            json={
                "objective_text": "整理工作量并输出 TXT",
                "upload_ids": [upload.upload_id],
                "output_formats": ["xlsx"],
            },
        )

        assert response.status_code == 422
        assert "要求输出 TXT" in response.json()["detail"]
        assert "界面选择了 XLSX" in response.json()["detail"]
        assert client.get("/api/semantic-workspace/tasks").json() == []


def test_workspace_explains_renderer_crash_without_publishing(
    tmp_path,
    monkeypatch,
):
    client, _ = _client(tmp_path, monkeypatch)
    upload = _upload(tmp_path)

    def crash_renderer(_path, _content):
        raise RuntimeError("模拟转换器崩溃")

    monkeypatch.setitem(
        delivery_service._RENDERERS,
        DeliveryFormat.XLSX,
        crash_renderer,
    )
    with client:
        created = client.post(
            "/api/semantic-workspace/tasks",
            json={
                "objective_text": (
                    "只提取谢超群的数据，只保留核销工作量天数和工作量费用"
                ),
                "upload_ids": [upload.upload_id],
                "output_formats": ["xlsx"],
            },
        )
        assert created.status_code == 202, created.text
        failed = _wait(
            client,
            created.json()["task_id"],
            {"failed"},
        )

    assert failed["failure"]["error_code"] == "DELIVERY_RENDER_FAILED", (
        json.dumps(
            {
                "failure": failed["failure"],
                "run": failed["run"],
                "attempts": failed["attempts"],
                "events": failed["harness_events"],
            },
            ensure_ascii=False,
            indent=2,
        )
    )
    assert failed["failure"]["stage"] == "deliver"
    assert failed["failure"]["source_read"] is True
    assert failed["failure"]["intermediate_created"] is True
    assert failed["failure"]["delivery_published"] is False
    assert failed["delivery"] is None
    assert not any(
        path.is_file()
        for path in (tmp_path / "executions").rglob("*.xlsx")
    )


def test_workspace_completes_four_real_file_delivery_loops(
    tmp_path,
    monkeypatch,
):
    client, _ = _client(
        tmp_path,
        monkeypatch,
        generator=Batch8AFileScenarioGenerator(),
    )
    fixture_root = (
        Path(__file__).parent
        / "fixtures"
        / "semantic_harness"
        / "public"
        / "batch0"
    )
    upload_store = UploadStore(
        root=str(tmp_path / "uploads"),
        max_bytes=10 * 1024 * 1024,
    )
    scenarios = (
        (
            "documents/contract.docx",
            "汇总这个 Word 里的商务条款",
            ["txt"],
        ),
        (
            "documents/contract.pdf",
            "核查违约责任并保留证据",
            ["docx", "pdf"],
        ),
        (
            "workload_filter/source.xlsx",
            "筛选排序示例人员甲的费用",
            ["xlsx"],
        ),
        (
            "workload_filter/source.csv",
            "按人员去重并分组汇总费用",
            ["csv", "json"],
        ),
    )

    with client:
        outputs: dict[str, bytes] = {}
        for relative_path, objective, formats in scenarios:
            fixture = fixture_root / relative_path
            uploaded = upload_store.save_bytes(
                "user-a",
                fixture.name,
                fixture.read_bytes(),
                media_type={
                    ".docx": (
                        "application/vnd.openxmlformats-officedocument."
                        "wordprocessingml.document"
                    ),
                    ".pdf": "application/pdf",
                    ".xlsx": (
                        "application/vnd.openxmlformats-officedocument."
                        "spreadsheetml.sheet"
                    ),
                    ".csv": "text/csv",
                }[fixture.suffix],
            )
            created = client.post(
                "/api/semantic-workspace/tasks",
                json={
                    "objective_text": objective,
                    "upload_ids": [uploaded.upload_id],
                    "output_formats": formats,
                    "provider": "local",
                    "model": "batch8a-file-scenarios",
                },
            )
            assert created.status_code == 202, created.text
            completed = _wait(
                client,
                created.json()["task_id"],
                {"completed", "failed", "needs_input"},
            )
            binding = (
                get_store().latest_semantic_binding_revision(
                    "user-a",
                    completed["plan_id"],
                )
                if completed["plan_id"]
                else None
            )
            assert completed["status"] == "completed", json.dumps(
                {
                    "task_status": completed["status"],
                    "task_error": completed["error"],
                    "run": completed["run"],
                    "attempts": completed["attempts"],
                    "harness_events": completed["harness_events"],
                    "binding_status": (
                        binding["status"] if binding else None
                    ),
                },
                ensure_ascii=False,
                indent=2,
            )
            assert {
                item["format"] for item in completed["delivery"]["outputs"]
            } == set(formats)
            assert all(
                item["qa"]["openable"]
                for item in completed["delivery"]["outputs"]
            )
            bundle = client.get(
                f"/api/semantic-workspace/tasks/{completed['task_id']}/bundle"
            )
            assert bundle.status_code == 200, bundle.text
            with zipfile.ZipFile(io.BytesIO(bundle.content)) as archive:
                for name in archive.namelist():
                    if name.startswith("outputs/"):
                        outputs[Path(name).suffix.lower()] = archive.read(name)

    txt = outputs[".txt"].decode("utf-8")
    assert "支付合同金额的百分之六十" in txt
    assert "九月三十日前完成全部成果交付" in txt
    assert "千分之一承担违约责任" in txt

    docx_text = "\n".join(
        paragraph.text
        for paragraph in Document(io.BytesIO(outputs[".docx"])).paragraphs
    )
    pdf_text = "\n".join(
        page.extract_text() or ""
        for page in PdfReader(io.BytesIO(outputs[".pdf"])).pages
    )
    assert "违约责任" in docx_text
    assert "违约责任" in pdf_text

    workbook = load_workbook(
        io.BytesIO(outputs[".xlsx"]),
        read_only=True,
        data_only=True,
    )
    xlsx_rows = list(workbook.active.iter_rows(values_only=True))
    workbook.close()
    assert xlsx_rows[0] == ("姓名", "工作量费用")
    assert all(row[0] == "示例人员甲" for row in xlsx_rows[1:])
    xlsx_costs = [float(row[1]) for row in xlsx_rows[1:]]
    assert xlsx_costs == sorted(xlsx_costs, reverse=True)

    csv_lines = outputs[".csv"].decode("utf-8-sig").splitlines()
    assert csv_lines[0] == "姓名,费用合计"
    json_rows = json.loads(outputs[".json"].decode("utf-8"))["rows"]
    with (fixture_root / "workload_filter/source.csv").open(
        "r",
        encoding="utf-8",
        newline="",
    ) as source:
        expected_names = {
            row["姓名"] for row in csv.DictReader(source)
        }
    assert {row["姓名"] for row in json_rows} == expected_names
    assert all(float(row["费用合计"]) > 0 for row in json_rows)


def test_workspace_clarification_keeps_scope_and_txt_excludes_other_clauses(
    tmp_path,
    monkeypatch,
) -> None:
    client, _ = _client(
        tmp_path,
        monkeypatch,
        generator=ScopedDocumentClarificationGenerator(),
    )
    upload = _upload_contract(tmp_path)
    with client:
        created = client.post(
            "/api/semantic-workspace/tasks",
            json={
                "objective_text": "这个合同里的付款条款是什么？汇总并输出 TXT",
                "upload_ids": [upload.upload_id],
                "output_formats": ["txt"],
            },
        )
        assert created.status_code == 202, created.text
        task_id = created.json()["task_id"]
        waiting = _wait(client, task_id, {"needs_input"})
        assert waiting["question"]["kind"] == "plan"
        assert waiting["question"]["question_id"] == "extract.mode"

        answered = client.post(
            f"/api/semantic-workspace/tasks/{task_id}/answer",
            json={"answer": "逐字原文"},
        )
        assert answered.status_code == 200, answered.text
        completed = _wait(client, task_id, {"completed", "failed"})
        assert completed["status"] == "completed", completed.get("error")

        plan_row = get_store().latest_semantic_plan_revision(
            "user-a",
            completed["plan_id"],
        )
        assert plan_row is not None
        assert plan_row["plan"]["source_scope"]["section_patterns"] == [
            "付款条款"
        ]
        assert plan_row["plan"]["record_grain"] == "条款"

        outputs = completed["delivery"]["outputs"]
        txt_output = next(item for item in outputs if item["format"] == "txt")
        record = get_store().get_semantic_delivery_output(
            "user-a",
            txt_output["output_id"],
        )
        assert record is not None
        text = Path(record["file_path"]).read_text(encoding="utf-8")
        assert "付款条款" in text
        assert "百分之六十" in text
        assert "交付条款" not in text
        assert "九月三十" not in text
        assert "违约责任" not in text


def test_workspace_external_gate_cancel_and_owner_isolation(
    tmp_path,
    monkeypatch,
):
    client, current_user = _client(tmp_path, monkeypatch)
    upload = _upload(tmp_path)
    with client:
        created = client.post(
            "/api/semantic-workspace/tasks",
            json={
                "objective_text": "整理工作量表",
                "upload_ids": [upload.upload_id],
                "output_formats": ["xlsx"],
                "provider": "deepseek",
            },
        )
        assert created.status_code == 202
        task_id = created.json()["task_id"]
        waiting = _wait(client, task_id, {"needs_input"})
        assert waiting["question"]["kind"] == "external"
        assert waiting["question"]["outbound_data"]

        current_user["value"] = "user-b"
        assert (
            client.get(
                f"/api/semantic-workspace/tasks/{task_id}"
            ).status_code
            == 404
        )
        assert client.get("/api/semantic-workspace/tasks").json() == []

        current_user["value"] = "user-a"
        cancelled = client.post(
            f"/api/semantic-workspace/tasks/{task_id}/cancel"
        )
        assert cancelled.status_code == 200
        assert cancelled.json()["status"] == "cancelled"
        assert get_store().latest_semantic_delivery("user-a", task_id) is None


def test_workspace_recycle_bin_is_user_scoped(
    tmp_path,
    monkeypatch,
):
    client, current_user = _client(tmp_path, monkeypatch)
    upload = _upload(tmp_path)
    with client:
        created = client.post(
            "/api/semantic-workspace/tasks",
            json={
                "objective_text": "整理工作量表",
                "upload_ids": [upload.upload_id],
                "provider": "deepseek",
            },
        ).json()
        task_id = created["task_id"]
        _wait(client, task_id, {"needs_input"})
        moved = client.delete(
            f"/api/semantic-workspace/tasks/{task_id}"
        )
        assert moved.status_code == 200
        assert moved.json()["deleted_at"]
        assert len(
            client.get(
                "/api/semantic-workspace/tasks?deleted=true"
            ).json()
        ) == 1

        current_user["value"] = "user-b"
        assert (
            client.get(
                "/api/semantic-workspace/tasks?deleted=true"
            ).json()
            == []
        )
        current_user["value"] = "user-a"
        restored = client.post(
            f"/api/semantic-workspace/tasks/{task_id}/restore"
        )
        assert restored.status_code == 200
        assert restored.json()["deleted_at"] is None


def test_workspace_revision_keeps_previous_delivery_available(
    tmp_path,
    monkeypatch,
):
    client, _ = _client(tmp_path, monkeypatch)
    upload = _upload(tmp_path)
    with client:
        created = client.post(
            "/api/semantic-workspace/tasks",
            json={
                "objective_text": (
                    "只提取谢超群的数据，只保留核销工作量天数和工作量费用"
                ),
                "upload_ids": [upload.upload_id],
                "output_formats": ["xlsx"],
            },
        )
        assert created.status_code == 202, created.text
        task_id = created.json()["task_id"]
        first = _wait(client, task_id, {"completed"})
        first_run_id = first["run_id"]

        revised = client.post(
            f"/api/semantic-workspace/tasks/{task_id}/revisions",
            json={
                "instruction": "保持筛选条件，同时增加 JSON 输出",
                "output_formats": ["xlsx", "json"],
            },
        )
        assert revised.status_code == 202, revised.text
        assert revised.json()["revision"] == 2
        second = _wait(client, task_id, {"completed"})
        assert second["viewing_revision"] == 2
        assert second["current_revision"] == 2
        assert second["run_id"] != first_run_id

        historical = client.get(
            f"/api/semantic-workspace/tasks/{task_id}?revision=1"
        )
        assert historical.status_code == 200, historical.text
        assert historical.json()["viewing_revision"] == 1
        assert historical.json()["current_revision"] == 2
        assert historical.json()["run_id"] == first_run_id
        assert historical.json()["delivery"]["outputs"]
        assert not any(
            event["event_type"] == "revision_created"
            for event in historical.json()["events"]
        )
        assert any(
            event["event_type"] == "revision_created"
            for event in second["events"]
        )

        preview = client.get(
            f"/api/semantic-workspace/tasks/{task_id}/preview?revision=1"
        )
        assert preview.status_code == 200, preview.text
        assert preview.json()["total"] == 11


def test_workspace_expired_recycle_records_are_removed(
    tmp_path,
    monkeypatch,
):
    client, _ = _client(tmp_path, monkeypatch)
    upload = _upload(tmp_path)
    with client:
        created = client.post(
            "/api/semantic-workspace/tasks",
            json={
                "objective_text": "整理工作量表",
                "upload_ids": [upload.upload_id],
                "provider": "deepseek",
            },
        ).json()
        task_id = created["task_id"]
        _wait(client, task_id, {"needs_input"})
        assert client.delete(
            f"/api/semantic-workspace/tasks/{task_id}"
        ).status_code == 200

        expired = datetime.now() - timedelta(seconds=1)
        get_store().update_semantic_workspace_task(
            "user-a",
            task_id,
            purge_after=expired.isoformat(timespec="seconds"),
        )
        assert (
            get_store().purge_expired_semantic_workspace_tasks(
                now=datetime.now()
            )
            == 1
        )
        assert (
            client.get(
                "/api/semantic-workspace/tasks?deleted=true"
            ).json()
            == []
        )
        tombstone = get_store().get_semantic_workspace_audit_tombstone(
            "user-a",
            task_id,
        )
        assert tombstone is not None
        assert tombstone["task_id"] == task_id
        assert tombstone["user_id"] == "user-a"
        assert len(tombstone["objective_sha256"]) == 64
        assert tombstone["source_refs"] == [{
            "upload_id": upload.upload_id,
            "sha256": upload.sha256,
        }]
        assert tombstone["result_refs"] == []
        assert tombstone["requested_formats"] == ["xlsx"]
        assert tombstone["terminal_status"] == "needs_input"
        assert tombstone["error_code"] is None
        assert tombstone["purge_reason"] == "retention_expired"
        serialized = json.dumps(tombstone, ensure_ascii=False)
        assert "整理工作量表" not in serialized
        assert "workload.csv" not in serialized
        assert str(tmp_path) not in serialized


@pytest.mark.asyncio
async def test_workspace_serializes_heavy_jobs(
    tmp_path,
    monkeypatch,
):
    monkeypatch.setattr(
        settings, "webui_db_path", str(tmp_path / "workspace.db")
    )
    monkeypatch.setattr(
        settings, "data_prep_upload_root", str(tmp_path / "uploads")
    )
    auth_mod._store = None
    upload = _upload(tmp_path)
    store = get_store()
    for index in range(2):
        store.create_semantic_workspace_task(
            "user-a",
            task_id=f"heavy-{index}",
            title=f"重任务 {index}",
            objective_text="生成正式 PDF",
            upload_ids=[upload.upload_id],
            output_formats=["pdf"],
            provider="local",
            model=None,
            external_api_confirmed=False,
        )

    manager = SemanticWorkspaceManager()
    active = 0
    peak = 0

    async def fake_run(_user_id: str, _task_id: str) -> None:
        nonlocal active, peak
        active += 1
        peak = max(peak, active)
        await asyncio.sleep(0.05)
        active -= 1

    monkeypatch.setattr(manager, "_run_task", fake_run)
    manager.start()
    try:
        await asyncio.wait_for(manager._queue.join(), timeout=3)
    finally:
        await manager.stop()
    assert peak == 1
