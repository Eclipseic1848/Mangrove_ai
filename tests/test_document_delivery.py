# -*- coding: utf-8 -*-
"""文档抽取权威交付测试。"""
from __future__ import annotations

import json
import hashlib

from openpyxl import load_workbook

from src.data_prep.artifact_store import ArtifactStore
from src.data_prep.document_models import (
    BoundingBox,
    DiscoverySpec,
    EvidenceRef,
    ExtractedField,
    ExtractedDocument,
    ExtractedRecord,
    ExtractedTable,
    ExtractionFieldSpec,
    ExtractionSpec,
    ExtractionStatus,
    ResultCardinality,
    ResultContract,
    ResultShape,
    TaskGoal,
)
from src.services.document_delivery import write_document_delivery


def test_document_delivery_writes_authoritative_manifest_and_quality(tmp_path):
    store = ArtifactStore(str(tmp_path))
    raw = store.write_raw(
        "doc-task",
        "upload:u1",
        b"contract",
        uri="contract.pdf",
        media_type="application/pdf",
        ext="pdf",
    )
    spec = ExtractionSpec(
        goal=TaskGoal(objective="提取合同编号"),
        discovery=DiscoverySpec(artifact_ids=[raw.artifact_id]),
        fields=[ExtractionFieldSpec(
            name="contract_no",
            required=True,
            description="合同编号",
        )],
    )
    evidence = EvidenceRef(
        artifact_id=raw.artifact_id,
        element_id="el-1",
        page=2,
        bbox=BoundingBox(
            x0=1, y0=1, x1=10, y1=10, coordinate_space="pdf_points"
        ),
        quote="contract_no: HT-001",
        quote_sha256=hashlib.sha256(
            "contract_no: HT-001".encode("utf-8")
        ).hexdigest(),
        extractor="pdfplumber",
        extractor_version="0.11.10",
        confidence=0.99,
    )
    field = ExtractedField(
        name="contract_no",
        value="HT-001",
        status=ExtractionStatus.FOUND,
        evidence_refs=[evidence],
    )

    delivery = write_document_delivery(
        store,
        "doc-task",
        spec=spec,
        raw_artifacts=[raw],
        fields=[field],
        review_tasks=[],
    )

    assert delivery.quality.overall.value == "pass"
    manifest = json.loads(
        (tmp_path / delivery.manifest_path).read_text(encoding="utf-8")
    )
    assert manifest["spec_version"] == "3"
    assert manifest["outputs"][0]["path"].endswith("extracted_fields.jsonl")
    assert manifest["outputs"][0]["records"] == 1
    assert manifest["outputs"][1]["format"] == "xlsx"
    assert manifest["outputs"][1]["records"] == 1
    assert manifest["quality_ref"].endswith("quality_report.json")
    assert manifest["lineage_ref"].endswith("extraction/lineage.jsonl")
    assert (tmp_path / "doc-task/extraction/evidence.jsonl").exists()
    xlsx_path = tmp_path / "doc-task/extraction/document_extraction.xlsx"
    workbook = load_workbook(xlsx_path, read_only=True)
    assert workbook.sheetnames == [
        "Fields",
        "Evidence",
        "Review",
        "Quality",
        "Manifest",
        "Rejects",
    ]
    assert workbook["Fields"].max_row == 2
    assert workbook["Evidence"].max_row == 2
    assert workbook["Manifest"]["B1"].value == "value"


def test_document_delivery_writes_record_sheet_and_one_sheet_per_table(tmp_path):
    store = ArtifactStore(str(tmp_path))
    raw = store.write_raw(
        "doc-task",
        "upload:u1",
        b"records",
        uri="records.docx",
        media_type=(
            "application/vnd.openxmlformats-officedocument."
            "wordprocessingml.document"
        ),
        ext="docx",
    )
    spec = ExtractionSpec(
        goal=TaskGoal(objective="全量提取工作和表格"),
        discovery=DiscoverySpec(artifact_ids=[raw.artifact_id]),
        fields=[ExtractionFieldSpec(name="工作内容")],
        result_contract=ResultContract(
            shape=ResultShape.RECORDS,
            cardinality=ResultCardinality.ALL,
            record_grain="一项工作",
            exhaustive=True,
        ),
    )
    evidence = EvidenceRef(
        artifact_id=raw.artifact_id,
        element_id="el-work",
        page=1,
        quote="张三负责需求分析",
        quote_sha256=hashlib.sha256(
            "张三负责需求分析".encode("utf-8")
        ).hexdigest(),
        extractor="python-docx",
        extractor_version="1.2.0",
        confidence=0.99,
        location={"kind": "docx_paragraph", "paragraph": 1},
    )
    record = ExtractedRecord(
        record_id="record-1",
        fields=[ExtractedField(
            name="工作内容",
            value="需求分析",
            status=ExtractionStatus.FOUND,
            evidence_refs=[evidence],
        )],
        source_artifact_ids=[raw.artifact_id],
    )
    tables = [
        ExtractedTable(
            table_id="table-1",
            name="人员表",
            artifact_id=raw.artifact_id,
            page=1,
            columns=["姓名", "金额"],
            rows=[{"姓名": "张三", "金额": 100}, {"姓名": "李四", "金额": 200}],
        ),
        ExtractedTable(
            table_id="table-2",
            name="任务表",
            artifact_id=raw.artifact_id,
            page=2,
            columns=["任务", "状态"],
            rows=[{"任务": "验收", "状态": "完成"}],
        ),
    ]

    delivery = write_document_delivery(
        store,
        "doc-task",
        spec=spec,
        raw_artifacts=[raw],
        fields=[],
        records=[record],
        tables=tables,
        review_tasks=[],
        coverage={"elements_processed": 4},
    )

    workbook = load_workbook(
        tmp_path / "doc-task/extraction/document_extraction.xlsx",
        read_only=True,
    )
    assert "Records" in workbook.sheetnames
    assert workbook["Records"].max_row == 2
    assert "人员表" in workbook.sheetnames
    assert "任务表" in workbook.sheetnames
    assert workbook["人员表"].max_row == 3
    manifest = json.loads(
        (tmp_path / delivery.manifest_path).read_text(encoding="utf-8")
    )
    paths = {item["path"] for item in manifest["outputs"]}
    assert any(path.endswith("extraction/extracted_records.jsonl") for path in paths)
    assert (tmp_path / "doc-task/extraction/extracted_tables.json").exists()


def test_document_delivery_fails_when_full_table_result_is_empty(tmp_path):
    store = ArtifactStore(str(tmp_path))
    raw = store.write_raw(
        "doc-empty-table",
        "upload:u1",
        b"no table",
        uri="notes.pdf",
        media_type="application/pdf",
        ext="pdf",
    )
    spec = ExtractionSpec(
        goal=TaskGoal(objective="提取所有表格"),
        discovery=DiscoverySpec(artifact_ids=[raw.artifact_id]),
        fields=[ExtractionFieldSpec(name="完整表格", required=True)],
        result_contract=ResultContract(
            shape=ResultShape.TABLES,
            cardinality=ResultCardinality.ALL,
            renderer="table_tabs",
            exhaustive=True,
        ),
    )

    delivery = write_document_delivery(
        store,
        "doc-empty-table",
        spec=spec,
        raw_artifacts=[raw],
        fields=[],
        tables=[],
        review_tasks=[],
        coverage={"elements_processed": 3, "table_rows": 0},
    )

    assert delivery.quality.overall.value == "fail"
    assert "任务要求全量输出，但未产出任何表格行" in delivery.quality.issues
    effective_dimension = next(
        item for item in delivery.quality.dimensions
        if item.name == "有效结果"
    )
    assert effective_dimension.passed is False


def test_document_delivery_accepts_table_json_as_authoritative_output(tmp_path):
    store = ArtifactStore(str(tmp_path))
    raw = store.write_raw(
        "doc-table-output",
        "upload:u1",
        b"table",
        uri="table.pdf",
        media_type="application/pdf",
        ext="pdf",
    )
    spec = ExtractionSpec(
        goal=TaskGoal(objective="提取完整表格"),
        discovery=DiscoverySpec(artifact_ids=[raw.artifact_id]),
        fields=[ExtractionFieldSpec(name="完整表格")],
        result_contract=ResultContract(
            shape=ResultShape.TABLES,
            cardinality=ResultCardinality.ALL,
            renderer="table_tabs",
            exhaustive=True,
        ),
    )
    table = ExtractedTable(
        table_id="table-1",
        name="表1",
        artifact_id=raw.artifact_id,
        page=1,
        columns=["列1", "列2"],
        rows=[{"列1": "A", "列2": 1}],
        evidence_element_ids=["element-1"],
    )

    delivery = write_document_delivery(
        store,
        "doc-table-output",
        spec=spec,
        raw_artifacts=[raw],
        fields=[],
        tables=[table],
        review_tasks=[],
        coverage={"elements_processed": 1, "table_rows": 1},
    )

    manifest = json.loads(
        (tmp_path / delivery.manifest_path).read_text(encoding="utf-8")
    )
    assert delivery.quality.overall.value == "pass"
    assert manifest["outputs"][0]["path"].endswith(
        "extraction/extracted_tables.json"
    )
    assert manifest["outputs"][0]["records"] == 1


def test_document_shape_writes_dedicated_authoritative_output(tmp_path):
    store = ArtifactStore(str(tmp_path))
    raw = store.write_raw(
        "doc-continuous",
        "upload:u1",
        b"document",
        uri="document.pdf",
        media_type="application/pdf",
        ext="pdf",
    )
    spec = ExtractionSpec(
        goal=TaskGoal(objective="输出连续文档"),
        discovery=DiscoverySpec(artifact_ids=[raw.artifact_id]),
        fields=[],
        result_contract=ResultContract(
            shape=ResultShape.DOCUMENT,
            renderer="document_view",
        ),
    )
    evidence = EvidenceRef(
        artifact_id=raw.artifact_id,
        element_id="el-doc",
        page=1,
        quote="连续正文",
        extractor="pdfplumber",
        extractor_version="0.11.10",
        confidence=0.99,
        location={"kind": "paragraph", "paragraph": 1},
    )
    document = ExtractedDocument(
        document_id="document-1",
        title="连续文档",
        content="连续正文",
        source_artifact_ids=[raw.artifact_id],
        evidence_refs=[evidence],
    )

    delivery = write_document_delivery(
        store,
        "doc-continuous",
        spec=spec,
        raw_artifacts=[raw],
        fields=[],
        documents=[document],
        review_tasks=[],
    )

    manifest = json.loads(
        (tmp_path / delivery.manifest_path).read_text(encoding="utf-8")
    )
    assert delivery.quality.overall.value == "pass"
    assert manifest["outputs"][0]["path"].endswith("extracted_documents.json")
    assert manifest["outputs"][0]["format"] == "json"
    workbook = load_workbook(
        tmp_path / "doc-continuous/extraction/document_extraction.xlsx",
        read_only=True,
    )
    assert "Documents" in workbook.sheetnames


def test_aggregate_shape_writes_dedicated_authoritative_output(tmp_path):
    store = ArtifactStore(str(tmp_path))
    raw = store.write_raw(
        "doc-aggregate",
        "upload:u1",
        b"aggregate",
        uri="aggregate.pdf",
        media_type="application/pdf",
        ext="pdf",
    )
    spec = ExtractionSpec(
        goal=TaskGoal(objective="汇总合同总额"),
        discovery=DiscoverySpec(artifact_ids=[raw.artifact_id]),
        fields=[ExtractionFieldSpec(name="合同总额", required=True)],
        result_contract=ResultContract(
            shape=ResultShape.AGGREGATE,
            renderer="aggregate_cards",
        ),
    )
    evidence = EvidenceRef(
        artifact_id=raw.artifact_id,
        element_id="el-total",
        page=1,
        quote="合同总额：100万元",
        extractor="pdfplumber",
        extractor_version="0.11.10",
        confidence=0.99,
        location={"kind": "paragraph", "paragraph": 1},
    )
    field = ExtractedField(
        name="合同总额",
        value="100万元",
        status=ExtractionStatus.FOUND,
        evidence_refs=[evidence],
    )

    delivery = write_document_delivery(
        store,
        "doc-aggregate",
        spec=spec,
        raw_artifacts=[raw],
        fields=[field],
        review_tasks=[],
    )

    manifest = json.loads(
        (tmp_path / delivery.manifest_path).read_text(encoding="utf-8")
    )
    assert delivery.quality.overall.value == "pass"
    assert manifest["outputs"][0]["path"].endswith("extracted_aggregates.json")
    payload = json.loads(
        (tmp_path / "doc-aggregate/extraction/extracted_aggregates.json")
        .read_text(encoding="utf-8")
    )
    assert payload[0]["values"] == {"合同总额": "100万元"}
    workbook = load_workbook(
        tmp_path / "doc-aggregate/extraction/document_extraction.xlsx",
        read_only=True,
    )
    assert "Aggregate" in workbook.sheetnames
