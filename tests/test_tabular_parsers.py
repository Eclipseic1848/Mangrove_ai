# -*- coding: utf-8 -*-
"""结构化表格解析器测试（Phase 2 Task 4）。"""
from __future__ import annotations

import io
from pathlib import Path

import pytest

from src.data_prep.artifact_store import ArtifactStore
from src.data_prep.models import RawArtifact
from src.parsers.tabular import TabularParser


def _make_csv_artifact(store: ArtifactStore, task_id: str, data: bytes, ext: str = "csv") -> RawArtifact:
    return store.write_raw(
        task_id=task_id, source_id="file-1", data=data,
        uri=f"upload://data.{ext}", media_type="text/csv", ext=ext,
    )


def test_csv_parser_parses_rows_with_position(tmp_path: Path):
    store = ArtifactStore(root=str(tmp_path))
    csv_bytes = b"id,name\n1,Alice\n2,Bob\n3,Charlie\n"
    art = _make_csv_artifact(store, "task-1", csv_bytes)

    parser = TabularParser()
    records, rejects = parser.parse(art, csv_bytes)

    assert len(records) == 3
    assert records[0].data["id"] == "1"
    assert records[0].data["name"] == "Alice"
    assert records[0].meta["position"]["row"] == 2
    assert records[0].meta["parser"] == "tabular_csv"
    assert records[0].meta["artifact_id"] == art.artifact_id
    assert rejects == []


def test_csv_parser_handles_utf8_bom(tmp_path: Path):
    store = ArtifactStore(root=str(tmp_path))
    csv_bytes = b"\xef\xbb\xbfid,name\n1,Alice\n"
    art = _make_csv_artifact(store, "task-1", csv_bytes)

    parser = TabularParser()
    records, _ = parser.parse(art, csv_bytes)

    assert len(records) == 1
    assert records[0].data["name"] == "Alice"


def test_csv_parser_handles_gbk(tmp_path: Path):
    store = ArtifactStore(root=str(tmp_path))
    csv_bytes = "id,name\n1,张三\n".encode("gbk")
    art = _make_csv_artifact(store, "task-1", csv_bytes)

    parser = TabularParser()
    records, _ = parser.parse(art, csv_bytes)

    assert len(records) == 1
    assert records[0].data["name"] == "张三"


def test_csv_parser_isolates_bad_row(tmp_path: Path):
    store = ArtifactStore(root=str(tmp_path))
    # 第 2 行字段数不一致（坏行），第 3 行正常
    csv_bytes = b"id,name\n1,Alice,EXTRA\n2,Bob\n"
    art = _make_csv_artifact(store, "task-1", csv_bytes)

    parser = TabularParser()
    records, rejects = parser.parse(art, csv_bytes)

    assert len(records) == 1
    assert records[0].data["name"] == "Bob"
    assert len(rejects) == 1
    assert rejects[0]["reason"] == "csv_bad_row"
    assert rejects[0]["position"]["row"] == 2


def test_csv_parse_stream_yields_batches(tmp_path: Path):
    store = ArtifactStore(root=str(tmp_path))
    rows = b"id,name\n" + b"".join(f"{i},name{i}\n".encode() for i in range(1, 6))
    art = _make_csv_artifact(store, "task-1", rows)

    parser = TabularParser()
    batches = list(parser.parse_stream(art, store.read_raw_bytes("task-1", art.storage_path), batch_size=2))

    assert len(batches) == 3  # 2 + 2 + 1
    all_records = [r for b in batches for r in b[0]]
    assert len(all_records) == 5
    assert all_records[0].meta["position"]["row"] == 2


def _make_xlsx_bytes(rows):
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    for row in rows:
        ws.append(row)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def test_excel_parser_parses_rows_with_position(tmp_path: Path):
    store = ArtifactStore(root=str(tmp_path))
    data = _make_xlsx_bytes([["id", "name"], [1, "Alice"], [2, "Bob"]])
    art = store.write_raw(
        task_id="task-1", source_id="file-1", data=data,
        uri="upload://data.xlsx",
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        ext="xlsx",
    )

    parser = TabularParser()
    records, rejects = parser.parse(art, data)

    assert len(records) == 2
    assert records[0].data["name"] == "Alice"
    assert records[0].meta["position"]["row"] == 2
    assert records[0].meta["parser"] == "tabular_xlsx"
    assert records[0].meta["artifact_id"] == art.artifact_id
    assert rejects == []


def test_excel_parser_handles_multiple_sheets(tmp_path: Path):
    from openpyxl import Workbook

    store = ArtifactStore(root=str(tmp_path))
    wb = Workbook()
    wb.active.title = "First"
    wb.active.append(["id", "name"])
    wb.active.append([1, "Alice"])
    ws2 = wb.create_sheet("Second")
    ws2.append(["id", "name"])
    ws2.append([2, "Bob"])
    buf = io.BytesIO()
    wb.save(buf)
    data = buf.getvalue()
    art = store.write_raw(
        task_id="task-1", source_id="file-1", data=data,
        uri="upload://multi.xlsx",
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        ext="xlsx",
    )

    parser = TabularParser()
    records, _ = parser.parse(art, data)

    assert len(records) == 2
    sheets = {r.meta["position"]["sheet"] for r in records}
    assert sheets == {"First", "Second"}


def _make_parquet_bytes(rows: dict):
    import pyarrow as pa
    import pyarrow.parquet as pq

    table = pa.table(rows)
    buf = io.BytesIO()
    pq.write_table(table, buf)
    return buf.getvalue()


def test_parquet_parser_parses_rows(tmp_path: Path):
    store = ArtifactStore(root=str(tmp_path))
    data = _make_parquet_bytes({"id": [1, 2, 3], "name": ["A", "B", "C"]})
    art = store.write_raw(
        task_id="task-1", source_id="file-1", data=data,
        uri="upload://data.parquet", media_type="application/vnd.apache.parquet", ext="parquet",
    )

    parser = TabularParser()
    records, rejects = parser.parse(art, data)

    assert len(records) == 3
    assert records[0].data["name"] == "A"
    assert records[0].meta["parser"] == "tabular_parquet"
    assert records[0].meta["artifact_id"] == art.artifact_id
    assert rejects == []
